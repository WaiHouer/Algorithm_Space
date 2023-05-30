"""
资源分配算法的入口，可以通过修改 tag 选择不同的算法
"""
import math
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from numba import njit
from Allocation_Epidemic_Function import allocation_epidemic_function
from Nature_Process import nature_process
from Myopic_LP import myopic_lp
from Myopic_Model import myopic_model
from Benchmark import benchmark
from ADP import adp
from GAQN import GAQN_Method
from GA_Basic_Function import coding
from VNS import Variable_Neighborhood_Search


np.set_printoptions(suppress=True)


class Resource_Allocation:
    def __init__(self):
        self.K = 9  # 地区数量
        self.T = 60  # 资源分配时间跨度
        self.sigma_hat = np.zeros((self.K, self.K))  # 区域之间的地理系数，矩阵
        '---------------------------------'
        self.book = load_workbook('Fitting_Result_v2.xlsx')
        self.group_initial = self.book[f'群体']
        self.para_initial = self.book[f'参数']
        self.info = self.book[f'区域信息']

        self.location = np.zeros((self.K, 2))  # 存放坐标
        self.N = np.zeros(self.K)  # 存放人口
        self.name = []  # 存放名字
        for i in range(self.K):
            self.N[i] = self.info.cell(2, i + 2).value
            self.name.append(self.info.cell(1, i + 2).value)
            for j in range(2):
                self.location[i][0] = self.info.cell(3, i + 2).value
                self.location[i][1] = self.info.cell(4, i + 2).value
        '---------------------------------'
        self.S_initial = np.zeros(self.K)  # 群体，初始条件
        self.E_initial = np.zeros(self.K)
        self.A_initial = np.zeros(self.K)
        self.Q_initial = np.zeros(self.K)
        self.U_initial = np.zeros(self.K)
        self.R_initial = np.zeros(self.K)
        self.D_initial = np.zeros(self.K)
        self.beta_e, self.beta_a, self.beta_u = np.zeros(self.K), np.zeros(self.K), np.zeros(self.K)  # 参数，初始条件
        self.alpha = np.zeros(self.K)
        self.delta_a, self.delta_q, self.delta_u = np.zeros(self.K), np.zeros(self.K), np.zeros(self.K)
        self.gamma_a, self.gamma_q, self.gamma_u = np.zeros(self.K), np.zeros(self.K), np.zeros(self.K)
        self.p, self.q = self.para_initial.cell(11, 2).value, self.para_initial.cell(12, 2).value
        self.sigma = self.para_initial.cell(13, 2).value
        for k in range(self.K):
            self.S_initial[k] = self.group_initial.cell(1, k + 2).value
            self.E_initial[k] = self.group_initial.cell(2, k + 2).value
            self.A_initial[k] = self.group_initial.cell(3, k + 2).value
            self.Q_initial[k] = self.group_initial.cell(4, k + 2).value
            self.U_initial[k] = self.group_initial.cell(5, k + 2).value
            self.R_initial[k] = self.group_initial.cell(6, k + 2).value
            self.D_initial[k] = self.group_initial.cell(7, k + 2).value
            self.beta_e[k] = self.para_initial.cell(1, k + 2).value
            self.beta_a[k] = self.para_initial.cell(2, k + 2).value
            self.beta_u[k] = self.para_initial.cell(3, k + 2).value
            self.alpha[k] = self.para_initial.cell(4, k + 2).value
            self.delta_a[k] = self.para_initial.cell(5, k + 2).value
            self.delta_q[k] = self.para_initial.cell(6, k + 2).value
            self.delta_u[k] = self.para_initial.cell(7, k + 2).value
            self.gamma_a[k] = self.para_initial.cell(8, k + 2).value
            self.gamma_q[k] = self.para_initial.cell(9, k + 2).value
            self.gamma_u[k] = self.para_initial.cell(10, k + 2).value
        self.sigma_hat = d_m_calculate(self.K, self.location, self.sigma)
        '---------------------------------'
        self.S_initial_last = np.zeros(self.K)  # 拟合的倒数第二个时期的群体，即：初始条件前一个时期
        self.E_initial_last = np.zeros(self.K)  # 用于计算U_new
        self.A_initial_last = np.zeros(self.K)
        self.Q_initial_last = np.zeros(self.K)
        self.U_initial_last = np.zeros(self.K)
        self.R_initial_last = np.zeros(self.K)
        self.D_initial_last = np.zeros(self.K)
        for k in range(self.K):
            self.S_initial_last[k] = self.group_initial.cell(11, k + 2).value
            self.E_initial_last[k] = self.group_initial.cell(12, k + 2).value
            self.A_initial_last[k] = self.group_initial.cell(13, k + 2).value
            self.Q_initial_last[k] = self.group_initial.cell(14, k + 2).value
            self.U_initial_last[k] = self.group_initial.cell(15, k + 2).value
            self.R_initial_last[k] = self.group_initial.cell(16, k + 2).value
            self.D_initial_last[k] = self.group_initial.cell(17, k + 2).value
        '---------------------------------'
        #  注：所有时间相关都用T+1，因为要把初期算上
        self.eta = 0.5  # 病床有效率
        self.b_hat = np.zeros(self.T + 1)  # 每期新增病床（注意别太大，容易问题不可行：b_last>U，累计病床过剩问题！！）
        for i in range(self.T + 1):
            self.b_hat[i] = 500
        self.lambda_b = 0.2  # 新增病床部署率（防止只发放给某一个区域）
        self.C = np.zeros(self.T + 1)  # 每期新增核酸检测（有几个区域系数为正，则不会分配给该区域，即使资源浪费）
        for i in range(self.T + 1):
            self.C[i] = 500000
        self.lambda_c = 0.2      # 新增核酸检测部署率（防止只发放给某一个区域）
        self.M = 50  # ADP算法，迭代修正次数
        self.L = 20  # ADP算法，单次修正，样本量
        self.O = 5 * self.K  # 状态变量（特征值）数量（没有参与任何计算，仅供展示）
        self.W = 20  # 神经元个数
        self.delta = 99999999  # 足够大的正数（没有参与任何计算，仅供展示）
        self.select_ratio = 0  # ADP算法被选择的轮盘赌，样本数上限= 此参数
        self.norm_tag = 'standard_norm'  # 归一化标准：None、'max_norm'、'standard_norm'
        '---------------------------------'

        self.algorithm()

    def algorithm(self):
        tag = 'all'

        value_nature = 0  # 自然状态下的总新增
        if tag:
            s_time = time.time()
            S, E, A, Q, U, R, D = \
                nature_process(self.K, self.T + 1, self.S_initial, self.E_initial, self.A_initial, self.Q_initial
                               , self.U_initial, self.R_initial, self.D_initial, self.N, self.sigma_hat, self.beta_e
                               , self.beta_a, self.beta_u, self.alpha, self.delta_a, self.delta_q, self.delta_u
                               , self.gamma_a, self.gamma_q, self.gamma_u, self.p, self.q, self.eta)
            e_time = time.time()
            value_nature_list = np.zeros(self.T + 1)
            value_nature_list_region = np.zeros((self.K, self.T + 1))
            for t in range(self.T + 1):
                for k in range(self.K):
                    value_nature_list[t] += E[k][t + 1] - E[k][t] + self.alpha[k] * E[k][t]
                    value_nature += E[k][t + 1] - E[k][t] + self.alpha[k] * E[k][t]
                    value_nature_list_region[k][t] = E[k][t + 1] - E[k][t] + self.alpha[k] * E[k][t]
            print(value_nature)
            print(f'自然状态迭代时间：{e_time - s_time}s')

        if tag == 'Myopic_LP' or tag == 'all':
            s_time = time.time()
            b_myopic, c_myopic, value_myopic, S_myopic, E_myopic, A_myopic, U_myopic, b_bar_myopic \
                = myopic_lp(self.K, self.T, self.S_initial, self.E_initial
                            , self.A_initial, self.Q_initial, self.U_initial
                            , self.R_initial, self.D_initial, self.N, self.sigma_hat
                            , self.beta_e, self.beta_a, self.beta_u, self.alpha
                            , self.delta_a, self.delta_q, self.delta_u
                            , self.gamma_a, self.gamma_q, self.gamma_u
                            , self.p, self.q, self.eta, self.b_hat, self.lambda_b
                            , self.C, self.lambda_c, re_tag=1)
            e_time = time.time()
            print(sum(value_myopic),'=>',f'减少新增：{value_nature - sum(value_myopic)}'
                                         f'，百分比：{(value_nature - sum(value_myopic)) / value_nature}')
            print(f'Myopic_LP算法运行时间：{e_time - s_time}s')

        if tag == 'Benchmark_Average' or tag == 'all':
            s_time = time.time()
            b_BH_Aver, c_BH_Aver, value_BH_Aver, S_BH_Aver, E_BH_Aver, A_BH_Aver, U_BH_Aver \
                = benchmark(self.K, self.T, self.S_initial, self.E_initial
                            , self.A_initial, self.Q_initial, self.U_initial
                            , self.R_initial, self.D_initial, self.N, self.sigma_hat
                            , self.beta_e, self.beta_a, self.beta_u, self.alpha
                            , self.delta_a, self.delta_q, self.delta_u
                            , self.gamma_a, self.gamma_q, self.gamma_u
                            , self.p, self.q, self.eta, self.b_hat, self.C
                            , 'Benchmark_Average', np.zeros(self.K), self.lambda_b, self.lambda_c, re_tag=1)
            e_time = time.time()
            print(sum(value_BH_Aver), '=>', f'减少新增：{value_nature - sum(value_BH_Aver)}'
                                            f'，百分比：{(value_nature - sum(value_BH_Aver)) / value_nature}')
            print(f'Benchmark_Average算法运行时间：{e_time - s_time}s')

        if tag == 'Benchmark_N' or tag == 'all':
            s_time = time.time()
            b_BH_N, c_BH_N, value_BH_N, S_BH_N, E_BH_N, A_BH_N, U_BH_N \
                = benchmark(self.K, self.T, self.S_initial, self.E_initial
                                                   , self.A_initial, self.Q_initial, self.U_initial
                                                   , self.R_initial, self.D_initial, self.N, self.sigma_hat
                                                   , self.beta_e, self.beta_a, self.beta_u, self.alpha
                                                   , self.delta_a, self.delta_q, self.delta_u
                                                   , self.gamma_a, self.gamma_q, self.gamma_u
                                                   , self.p, self.q, self.eta, self.b_hat, self.C
                                                   , 'Benchmark_N', np.zeros(self.K), self.lambda_b, self.lambda_c, re_tag=1)
            e_time = time.time()
            print(sum(value_BH_N), '=>', f'减少新增：{value_nature - sum(value_BH_N)}'
                                         f'，百分比：{(value_nature - sum(value_BH_N)) / value_nature}')
            print(f'Benchmark_N算法运行时间：{e_time - s_time}s')

        if tag == 'Benchmark_U' or tag == 'all':
            s_time = time.time()
            b_BH_U, c_BH_U, value_BH_U, S_BH_U, E_BH_U, A_BH_U, U_BH_U \
                = benchmark(self.K, self.T, self.S_initial, self.E_initial
                                                   , self.A_initial, self.Q_initial, self.U_initial
                                                   , self.R_initial, self.D_initial, self.N, self.sigma_hat
                                                   , self.beta_e, self.beta_a, self.beta_u, self.alpha
                                                   , self.delta_a, self.delta_q, self.delta_u
                                                   , self.gamma_a, self.gamma_q, self.gamma_u
                                                   , self.p, self.q, self.eta, self.b_hat, self.C
                                                   , 'Benchmark_U', np.zeros(self.K), self.lambda_b, self.lambda_c, re_tag=1)
            e_time = time.time()
            print(sum(value_BH_U), '=>', f'减少新增：{value_nature - sum(value_BH_U)}'
                                         f'，百分比：{(value_nature - sum(value_BH_U)) / value_nature}')
            print(f'Benchmark_U算法运行时间：{e_time - s_time}s')

        if tag == 'Benchmark_U_new' or tag == 'all':
            s_time = time.time()
            b_BH_U_n, c_BH_U_n, value_BH_U_n, S_BH_U_n, E_BH_U_n, A_BH_U_n, U_BH_U_n \
                = benchmark(self.K, self.T, self.S_initial, self.E_initial
                                                         , self.A_initial, self.Q_initial, self.U_initial
                                                         , self.R_initial, self.D_initial, self.N, self.sigma_hat
                                                         , self.beta_e, self.beta_a, self.beta_u, self.alpha
                                                         , self.delta_a, self.delta_q, self.delta_u
                                                         , self.gamma_a, self.gamma_q, self.gamma_u
                                                         , self.p, self.q, self.eta, self.b_hat, self.C
                                                         , 'Benchmark_U_new', np.zeros(self.K), self.lambda_b, self.lambda_c, self.E_initial_last, re_tag=1)
            e_time = time.time()
            print(sum(value_BH_U_n), '=>', f'减少新增：{value_nature - sum(value_BH_U_n)}'
                                           f'，百分比：{(value_nature - sum(value_BH_U_n)) / value_nature}')
            print(f'Benchmark_U_new算法运行时间：{e_time - s_time}s')

        if tag == 'allaa':
            s_time = time.time()
            b_ADP, c_ADP, value_ADP = adp(self.K, self.T, self.S_initial, self.E_initial
                                          , self.A_initial, self.Q_initial, self.U_initial
                                          , self.R_initial, self.D_initial, self.N, self.sigma_hat
                                          , self.beta_e, self.beta_a, self.beta_u, self.alpha
                                          , self.delta_a, self.delta_q, self.delta_u
                                          , self.gamma_a, self.gamma_q, self.gamma_u
                                          , self.p, self.q, self.eta, self.b_hat, self.C, self.E_initial_last
                                          , self.M, self.L, self.lambda_b, self.lambda_c, self.W
                                          , S_myopic, E_myopic, A_myopic, U_myopic, b_myopic
                                          , S_BH_Aver, E_BH_Aver, A_BH_Aver, U_BH_Aver, b_BH_Aver
                                          , S_BH_N, E_BH_N, A_BH_N, U_BH_N, b_BH_N
                                          , S_BH_U, E_BH_U, A_BH_U, U_BH_U, b_BH_U
                                          , S_BH_U_n, E_BH_U_n, A_BH_U_n, U_BH_U_n, b_BH_U_n
                                          , self.select_ratio, norm_tag=self.norm_tag)
            e_time = time.time()
            print(sum(value_ADP), '=>', f'减少新增：{value_nature - sum(value_ADP)}'
                                        f'，百分比：{(value_nature - sum(value_ADP)) / value_nature}')
            print(f'ADP算法运行时间：{e_time - s_time}s')
            print(b_ADP, c_ADP)

        if tag == 'allaa':
            s_time = time.time()
            b_ADP_Short, c_ADP_Short = np.zeros((self.K, self.T + 1)), np.zeros((self.K, self.T + 1))
            value_ADP_Short = np.zeros(self.T + 1)

            T_short, T_tem = 2, 0
            S_initial, E_initial, A_initial = self.S_initial, self.E_initial, self.A_initial
            Q_initial, U_initial, R_initial, D_initial = self.Q_initial, self.U_initial, self.R_initial, self.D_initial
            E_initial_last = self.E_initial_last
            b_before = np.zeros(self.K)
            B_add = 0  # 因为短视ADP会遗忘前面的累计病床，需要另外给加上，否则不可行
            while T_tem <= self.T:
                if T_tem + T_short <= self.T:
                    b_, c_, value_, S_, E_, A_, Q_, U_, R_, D_\
                        = adp(self.K, T_short - 1, S_initial, E_initial, A_initial, Q_initial
                              , U_initial, R_initial, D_initial, self.N, self.sigma_hat
                              , self.beta_e, self.beta_a, self.beta_u, self.alpha, self.delta_a, self.delta_q
                              , self.delta_u, self.gamma_a, self.gamma_q, self.gamma_u, self.p, self.q, self.eta
                              , self.b_hat[T_tem:T_tem + T_short], self.C[T_tem:T_tem + T_short]
                              , E_initial_last, self.M, self.L, self.lambda_b, self.lambda_c, self.W
                              , S_myopic[:, T_tem:T_tem + T_short], E_myopic[:, T_tem:T_tem + T_short]
                              , A_myopic[:, T_tem:T_tem + T_short], U_myopic[:, T_tem:T_tem + T_short]
                              , b_myopic[:, T_tem:T_tem + T_short]
                              , S_BH_Aver[:, T_tem:T_tem + T_short], E_BH_Aver[:, T_tem:T_tem + T_short]
                              , A_BH_Aver[:, T_tem:T_tem + T_short], U_BH_Aver[:, T_tem:T_tem + T_short]
                              , b_BH_Aver[:, T_tem:T_tem + T_short]
                              , S_BH_N[:, T_tem:T_tem + T_short], E_BH_N[:, T_tem:T_tem + T_short]
                              , A_BH_N[:, T_tem:T_tem + T_short], U_BH_N[:, T_tem:T_tem + T_short]
                              , b_BH_N[:, T_tem:T_tem + T_short]
                              , S_BH_U[:, T_tem:T_tem + T_short], E_BH_U[:, T_tem:T_tem + T_short]
                              , A_BH_U[:, T_tem:T_tem + T_short], U_BH_U[:, T_tem:T_tem + T_short]
                              , b_BH_U[:, T_tem:T_tem + T_short]
                              , S_BH_U_n[:, T_tem:T_tem + T_short], E_BH_U_n[:, T_tem:T_tem + T_short]
                              , A_BH_U_n[:, T_tem:T_tem + T_short], U_BH_U_n[:, T_tem:T_tem + T_short]
                              , b_BH_U_n[:, T_tem:T_tem + T_short]
                              , self.select_ratio, re_tag=1, b_before=b_before, B_add=B_add
                              , norm_tag=self.norm_tag)

                    b_ADP_Short[:, T_tem:T_tem + T_short] = b_[:, :]
                    c_ADP_Short[:, T_tem:T_tem + T_short] = c_[:, :]
                    value_ADP_Short[T_tem:T_tem + T_short] = value_[:]

                    # 更新状态
                    S_nxt, E_nxt, A_nxt, Q_nxt, U_nxt, R_nxt, D_nxt \
                        = allocation_epidemic_function(self.K, 1, S_[:, -1], E_[:, -1], A_[:, -1], Q_[:, -1], U_[:, -1]
                                                       , R_[:, -1], D_[:, -1], self.N, self.sigma_hat
                                                       , self.beta_e, self.beta_a, self.beta_u, self.alpha
                                                       , self.delta_a, self.delta_q, self.delta_u
                                                       , self.gamma_a, self.gamma_q, self.gamma_u
                                                       , self.p, self.q, b_[:, -1].reshape(self.K, 1)
                                                       , c_[:, -1].reshape(self.K, 1), self.eta)

                    S_initial = S_nxt[:, -1]
                    E_initial = E_nxt[:, -1]
                    A_initial = A_nxt[:, -1]
                    Q_initial = Q_nxt[:, -1]
                    U_initial = U_nxt[:, -1]
                    R_initial = R_nxt[:, -1]
                    D_initial = D_nxt[:, -1]

                    E_initial_last = E_[:, -1]  # 这里注意一下
                    b_before = b_[:, -1]
                    B_add += sum(self.b_hat[T_tem:T_tem + T_short])  # 因为短视ADP会遗忘前面的累计病床，需要另外给加上，否则不可行
                    print(f'第{T_tem}~{T_tem + T_short - 1}期，完成')

                elif T_tem == self.T:  # 最后只剩下一期，则myopic
                    b_, c_, value_\
                        = myopic_model(self.K, S_initial, E_initial, A_initial, U_initial, b_before, self.N
                                       , self.sigma_hat, self.beta_e, self.beta_a, self.beta_u, self.eta
                                       , self.b_hat[T_tem:], self.lambda_b, self.C[T_tem:], self.lambda_c, B_add=B_add)

                    b_ADP_Short[:, -1] = b_[:]
                    c_ADP_Short[:, -1] = c_[:]
                    value_ADP_Short[-1] = value_
                else:  # 最后剩下超过一期，但是不足短视周期
                    b_, c_, value_, S_, E_, A_, Q_, U_, R_, D_ \
                        = adp(self.K, self.T - T_tem, S_initial, E_initial, A_initial, Q_initial
                              , U_initial, R_initial, D_initial, self.N, self.sigma_hat
                              , self.beta_e, self.beta_a, self.beta_u, self.alpha, self.delta_a, self.delta_q
                              , self.delta_u, self.gamma_a, self.gamma_q, self.gamma_u, self.p, self.q, self.eta
                              , self.b_hat[T_tem:], self.C[T_tem:]
                              , E_initial_last, self.M, self.L, self.lambda_b, self.lambda_c, self.W
                              , S_myopic[:, T_tem:], E_myopic[:, T_tem:]
                              , A_myopic[:, T_tem:], U_myopic[:, T_tem:]
                              , b_myopic[:, T_tem:]
                              , S_BH_Aver[:, T_tem:], E_BH_Aver[:, T_tem:]
                              , A_BH_Aver[:, T_tem:], U_BH_Aver[:, T_tem:]
                              , b_BH_Aver[:, T_tem:]
                              , S_BH_N[:, T_tem:], E_BH_N[:, T_tem:]
                              , A_BH_N[:, T_tem:], U_BH_N[:, T_tem:]
                              , b_BH_N[:, T_tem:]
                              , S_BH_U[:, T_tem:], E_BH_U[:, T_tem:]
                              , A_BH_U[:, T_tem:], U_BH_U[:, T_tem:]
                              , b_BH_U[:, T_tem:]
                              , S_BH_U_n[:, T_tem:], E_BH_U_n[:, T_tem:]
                              , A_BH_U_n[:, T_tem:], U_BH_U_n[:, T_tem:]
                              , b_BH_U_n[:, T_tem:]
                              , self.select_ratio, re_tag=1, b_before=b_before, B_add=B_add
                              , norm_tag=self.norm_tag)

                    b_ADP_Short[:, T_tem:] = b_[:, :]
                    c_ADP_Short[:, T_tem:] = c_[:, :]
                    value_ADP_Short[T_tem:] = value_[:]

                T_tem += T_short

            e_time = time.time()
            print(sum(value_ADP_Short), '=>', f'减少新增：{value_nature - sum(value_ADP_Short)}'
                                        f'，百分比：{(value_nature - sum(value_ADP_Short)) / value_nature}')
            print(f'短视ADP算法运行时间：{e_time - s_time}s')
            print(b_ADP_Short, c_ADP_Short)

        if tag == 'GAQN' or tag == 'allaa':
            s_time = time.time()
            GAQN_Method(self.K, self.T, self.S_initial, self.E_initial, self.A_initial, self.Q_initial, self.U_initial
                        , self.R_initial, self.D_initial, self.N, self.sigma_hat, self.beta_e, self.beta_a
                        , self.beta_u, self.alpha, self.delta_a, self.delta_q, self.delta_u, self.gamma_a
                        , self.gamma_q, self.gamma_u, self.p, self.q, self.eta, self.b_hat, self.lambda_b
                        , self.C, self.lambda_c, b_myopic, c_myopic, b_BH_Aver, c_BH_Aver, b_BH_N, c_BH_N
                        , b_BH_U, c_BH_U, b_BH_U_n, c_BH_U_n
                        , value_nature_list, value_nature_list_region)
            e_time = time.time()
            print(f'GA算法运行时间：{e_time - s_time}s')

        if tag == 'VNS' or tag == 'all':
            s_time = time.time()
            b_tem = coding(b_bar_myopic)  # 因为，VNS要传入编码后的b资源（分配量）
            B_last = np.zeros(self.K)
            VNS_result = Variable_Neighborhood_Search(self.K, self.S_initial, self.E_initial, self.A_initial
                                                      , self.Q_initial, self.U_initial, self.R_initial, self.D_initial
                                                      , self.N, self.sigma_hat, self.beta_e, self.beta_a, self.beta_u
                                                      , self.alpha, self.delta_a, self.delta_q, self.delta_u
                                                      , self.gamma_a, self.gamma_q, self.gamma_u, self.p, self.q
                                                      , self.eta, value_nature_list_region, b_tem, c_myopic, B_last
                                                      , self.b_hat, self.C, self.lambda_b, self.lambda_c, b_myopic)
            e_time = time.time()
            # print(VNS_result.b_final)
            # print(VNS_result.c_final)
            print(value_nature - VNS_result.fitness_final, '=>', f'减少新增：{VNS_result.fitness_final}'
                                                           f'，百分比：{VNS_result.fitness_final / value_nature}')
            print(f'VNS算法运行时间：{e_time - s_time}s')


@njit()
def d_m_calculate(K, location, c_0):  # 计算m值
    m = np.zeros((K, K))
    d = np.zeros((K, K))
    for k in range(K):
        for j in range(K):  # 计算地理距离的倒数
            if k != j:
                d[k][j] = 1 / math.sqrt((location[k][0] - location[j][0]) ** 2 +
                                        (location[k][1] - location[j][1]) ** 2)
    for k in range(K):
        for j in range(K):  # 计算地理系数
            if k == j:
                m[k][j] = c_0
            else:
                m[k][j] = (1 - c_0) * d[k][j] / sum(d[:, j])
    return m


if __name__ == '__main__':
    Resource_Allocation()

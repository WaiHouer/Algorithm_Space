"""
多区域—SEAQURD传染病模型
"""
from Region import Region
import numpy as np
import math
from openpyxl import load_workbook
import matplotlib.pyplot as plt


class SEAQURD:
    def __init__(self,region_num,file_name,start,end,total_population,S,E,A,Q,U,R,D,para):
        self.book = load_workbook(file_name)
        self.sheet = []
        self.region_num = region_num
        for i in range(self.region_num):
            self.sheet.append(self.book[f'{i}'])  # 将每一个sheet存入列表

        self.start = start  # 开始时间点
        self.end = end  # 结束时间点
        self.T = self.end - self.start + 1  # 时间长度

        self.total_population = total_population  # 各州总人口数

        self.region = []  # 用于存放地区

        # 参数列表：[[beta_e],[beta_a],[beta_u],[alpha],[delta_a],[delta_q],[delta_u],[gamma_a],[gamma_q],[gamma_u],p,q,c_0]
        self.para = para

        self.actual = [[] for i in range(self.region_num)]
        for i in range(self.region_num):
            for j in range(self.T):
                self.actual[i].append(self.sheet[i].cell(1,self.start+j+89).value)
        print(self.actual)
        print('读取数据完毕')

        # # 用于存放参数
        # self.beta_e, self.beta_a, self.beta_u = [], [], []
        # self.alpha = []
        # self.delta_a, self.delta_q, self.delta_u = [], [], []
        # self.gamma_a, self.gamma_q, self.gamma_u = [], [], []

        self.beta_e, self.beta_a, self.beta_u, self.alpha, self.delta_a, self.delta_q, self.delta_u, \
        self.gamma_a, self.gamma_q, self.gamma_u, self.p, self.q, self.c_0 = self.para

        # self.p = 0.3  # 群体I中转化为A的占比
        # self.q = 0.75  # 群体Y中转化为Q的占比
        # self.c_0 = 0.8  # 本地的地理相关接触率系数

        for i in range(self.region_num):
            self.region.append(Region(f'{self.sheet[i].cell(1,1).value}',
                                      [self.sheet[i].cell(1,4).value,self.sheet[i].cell(1,5).value],
                                      self.total_population[i],
                                      S[i],E[i],A[i],Q[i],U[i],R[i],D[i]))

        # self.region.append(Region('0-region',[1,45],10000,9900,100,0,0,0,0,0))
        # self.beta_e.append(0.21), self.beta_a.append(0.35), self.beta_u.append(0.4)
        # self.alpha.append(0.1)
        # self.delta_a.append(0.0001), self.delta_q.append(0.0001), self.delta_u.append(0.0002)
        # self.gamma_a.append(0.013), self.gamma_q.append(0.015), self.gamma_u.append(0.01)

        # self.region.append(Region('1-region',[3,7],8900,8800,100,0,0,0,0,0))
        # self.beta_e.append(0.23), self.beta_a.append(0.30), self.beta_u.append(0.32)
        # self.alpha.append(0.1)
        # self.delta_a.append(0.0001), self.delta_q.append(0.0001), self.delta_u.append(0.0002)
        # self.gamma_a.append(0.013), self.gamma_q.append(0.015), self.gamma_u.append(0.01)
        #
        # self.region.append(Region('2-region',[11,30],9500,9400,100,0,0,0,0,0))
        # self.beta_e.append(0.24), self.beta_a.append(0.32), self.beta_u.append(0.35)
        # self.alpha.append(0.1)
        # self.delta_a.append(0.0001), self.delta_q.append(0.0001), self.delta_u.append(0.0002)
        # self.gamma_a.append(0.013), self.gamma_q.append(0.015), self.gamma_u.append(0.01)

        self.m = np.zeros((len(self.region),len(self.region)))  # 初始化距离矩阵

        # 每行对应一个地区，列为时间间隔
        self.S = np.zeros((len(self.region), self.T))
        self.E = np.zeros((len(self.region), self.T))
        self.A = np.zeros((len(self.region), self.T))
        self.Q = np.zeros((len(self.region), self.T))
        self.U = np.zeros((len(self.region), self.T))
        self.R = np.zeros((len(self.region), self.T))
        self.D = np.zeros((len(self.region), self.T))
        for i in range(len(self.region)):
            self.S[i][0] = self.region[i].S
            self.E[i][0] = self.region[i].E
            self.A[i][0] = self.region[i].A
            self.Q[i][0] = self.region[i].Q
            self.U[i][0] = self.region[i].U
            self.R[i][0] = self.region[i].R
            self.D[i][0] = self.region[i].D

        self.I = np.zeros((len(self.region), self.T))  # 即：所有感染者数量 = A + Q + U

        # 两种生成距离倒数矩阵的方式：1：按照坐标计算，2：直接指定
        self.dist = self.dist_type_1()
        # self.dist = self.dist_type_2()

        self.algorithm()

        # self.picture()

    def algorithm(self):
        self.m_calculate()

        for i in range(len(self.region)):

            for t in range(1,self.T):

                e_rate = 0
                a_rate = 0
                u_rate = 0
                for j in range(len(self.region)):
                    e_rate += self.E[j][t-1] / self.region[j].N * self.beta_e[j] * self.m[i][j]
                    a_rate += self.A[j][t-1] / self.region[j].N * self.beta_a[j] * self.m[i][j]
                    u_rate += self.U[j][t-1] / self.region[j].N * self.beta_u[j] * self.m[i][j]

                self.S[i][t] = self.S[i][t-1] - self.S[i][t-1] * (e_rate + a_rate + u_rate)
                self.E[i][t] = self.E[i][t-1] + self.S[i][t-1] * (e_rate + a_rate + u_rate) \
                               - self.alpha[i] * self.E[i][t-1]
                self.A[i][t] = self.A[i][t-1] + self.p * self.alpha[i] * self.E[i][t-1] \
                               - self.delta_a[i] * self.A[i][t-1] - self.gamma_a[i] * self.A[i][t-1]
                self.Q[i][t] = self.Q[i][t-1] + (1 - self.p) * self.q * self.alpha[i] * self.E[i][t-1] \
                               - self.delta_q[i] * self.Q[i][t-1] - self.gamma_q[i] * self.Q[i][t-1]
                self.U[i][t] = self.U[i][t-1] + (1 - self.p) * (1- self.q) * self.alpha[i] * self.E[i][t-1] \
                               - self.delta_u[i] * self.U[i][t-1] - self.gamma_u[i] * self.U[i][t-1]
                self.R[i][t] = self.R[i][t-1] + self.gamma_a[i] * self.A[i][t-1] \
                               + self.gamma_q[i] * self.Q[i][t-1] \
                               + self.gamma_u[i] * self.U[i][t-1]
                self.D[i][t] = self.D[i][t-1] + self.delta_a[i] * self.A[i][t-1] \
                               + self.delta_q[i] * self.Q[i][t-1] \
                               + self.delta_u[i] * self.U[i][t-1]
        self.I = self.A + self.Q +self.U

    def m_calculate(self):
        for i in range(len(self.region)):
            for j in range(len(self.region)):
                if i == j:
                    self.m[i][j] = self.c_0
                else:
                    self.m[i][j] = (1 - self.c_0) * self.dist[i][j] / sum(self.dist[:,j])

    def dist_type_1(self):
        dist = np.zeros((len(self.region),len(self.region)))

        for i in range(len(self.region)):
            for j in range(len(self.region)):
                if i != j:
                    dist[i][j] = 1 / math.sqrt((self.region[i].location[0] - self.region[j].location[0]) ** 2 +
                                               (self.region[i].location[1] - self.region[j].location[1]) ** 2)

        return dist

    def dist_type_2(self):
        dist = np.zeros((len(self.region),len(self.region)))

        dist[0][1] = dist[1][0] = 1 / 38
        dist[0][2] = dist[2][0] = 1 / 18
        dist[1][2] = dist[2][1] = 1 / 24

        return dist

    def picture(self):
        t_range = np.arange(0, self.T)  # 时间跨度，分成一天份

        for i in range(len(self.region)):
            plt.subplot(2,3,i+1)

            # plt.plot(t_range, self.S[i], color='darkblue', label='S', marker='.')  # 画出易感者
            plt.plot(t_range, self.E[i], color='orange', label='E', marker='.')  # 画出潜伏着
            plt.plot(t_range, self.A[i], color='red', label='A', marker='.')  # 画出无症状感染者
            plt.plot(t_range, self.Q[i], color='purple', label='Q', marker='.')  # 画出有症状-隔离者
            plt.plot(t_range, self.U[i], color='olivedrab', label='U', marker='.')  # 画出有症状-未隔离者
            plt.plot(t_range, self.R[i], color='green', label='R', marker='.')  # 画出治愈者
            plt.plot(t_range, self.D[i], color='black', label='D', marker='.')  # 画出死亡者

            plt.plot(t_range, self.I[i], color='darkred', label='I = A + Q + U', marker='.')  # 画出死亡者

            plt.scatter(t_range, self.actual[i], color='lightblue', label='actual num')  # 画出真实感染数量

            plt.title(f'{self.region[i].name}')
            plt.legend(fontsize=10, facecolor='lightyellow')
            plt.xlabel('Day')
            plt.ylabel('Number')
        plt.show()

        total_I = []
        total_actual = []
        for i in range(self.T):
            ii = 0
            aa = 0
            for j in range(self.region_num):
                ii += self.I[j][i]
                aa += self.actual[j][i]
            total_I.append(ii)
            total_actual.append(aa)
        plt.plot(t_range, total_I, label='total_I')
        plt.plot(t_range, total_actual, label='total_actual')
        plt.legend(fontsize=10, facecolor='lightyellow')
        plt.show()


if __name__ == '__main__':
    SEAQURD(6,'American_data.xlsx',0,0+60-1,
            [4908621,734002,7378494,3038999,39937489,5845526],
            [4908621-3870,734002-281,7378494-3705,3038999-1374,39937489-23403,5845526-7551],
            [0,0,0,0,0,0],
            [0,0,0,0,0,0],
            [0,0,0,0,0,0],
            [3870,281,3705,1374,23403,7551],
            [0,0,0,0,0,0],
            [0,0,0,0,0,0],
            [[0.05,0.06,0.065,0.07,0.06,0.055],[0.055,0.06,0.05,0.075,0.065,0.06],[0.06,0.065,0.06,0.075,0.065,0.065],
             [0.1,0.1,0.1,0.1,0.1,0.1],[0.0001,0.0001,0.0001,0.0001,0.0001,0.0001],[0.0001,0.0001,0.0001,0.0001,0.0001,0.0001],
             [0.0002,0.0002,0.0002,0.0002,0.0002,0.0002],[0.013,0.013,0.013,0.013,0.013,0.013],[0.015,0.015,0.015,0.015,0.015,0.015],
             [0.010,0.010,0.010,0.010,0.010,0.010],0.3,0.8,0.9])

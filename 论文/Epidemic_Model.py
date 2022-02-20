"""
完整的传染病模型入口，包括：模型，参数拟合，多峰判断
"""
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
from Metropolis_Hastings import Metropolis_Hastings
from SEIR import SEIR
from Multipeak_judge import Multipeak_judge


class Epidemic_Model:  # 完整传染病模型
    def __init__(self,file_name):
        self.file_name = file_name
        self.book = load_workbook(file_name)
        self.sheet = self.book['湖北']

        # 记录完整的拟合区间（如：前60天）
        self.start = 0  # 开始时间点
        self.end = 59  # 结束时间点
        self.t_num = self.end - self.start + 1  # 时间长度

        self.actual = []  # 真实的感染人数，读取数据即可
        for i in range(self.start,self.end+1):
            self.actual.append(self.sheet.cell(4+i,4).value)

        # 初始化群体，用于记录“最终”的拟合结果（即：整合了多峰后的最终结果）
        self.S = [0 for i in range(self.t_num)]
        self.E = [0 for i in range(self.t_num)]
        self.I = [0 for i in range(self.t_num)]
        self.R = [0 for i in range(self.t_num)]

        # 总人数（不变）
        self.total_population = 150000

        self.new_flow_node = []  # 存放新浪潮开始节点（方便画图）

        self.algorithm()  # 拟合总算法
        self.picture()

    def algorithm(self):  # 拟合总算法
        # 初始化各群体数量
        S_0 = self.total_population-41
        E_0 = 0
        I_0 = 41
        R_0 = 0

        start = self.start  # 初始化拟合起点（随迭代向后推动变化）
        end = self.end  # 初始化拟合终点（始终不变）

        while True:  # 循环：不断迭代，对新浪潮进行拟合
            # （1）MCMC算法，参数拟合
            # 输入：文件名，起点，终点，总人数，各群体初值
            # 拟合后得到：para，为参数的集合列表
            sample = Metropolis_Hastings(self.file_name,start,end,self.total_population,S_0,E_0,I_0,R_0)
            # print(sample.para)

            # （2）将拟合得到的参数，输入传染病模型，从而得到各群体的拟合数量
            # 输入：文件名，起点，终点，总人数，各群体初值，拟合好的参数
            # 模型运算后得到：各群体拟合数量
            # 注：拟合数量的list长度，随着起点的向后推移而逐渐变短
            fitting = SEIR(self.file_name,start,end,self.total_population,S_0,E_0,I_0,R_0,sample.para)

            # 将真实感染人数list进行切片，目的是与拟合list长度和对应区间保持一致
            # 同样也是随着起点向后变短
            act = self.actual[start:end+1]
            # print(len(act),len(fitting.I))

            # （3）多峰判断，确定变化点
            # 输入：真实感染人数list（切片后），该区间的拟合好的S数量，该区间的拟合好的I数量
            # 得到：新浪潮开始节点peak_node
            peak = Multipeak_judge(act,fitting.S,fitting.I)

            for i in range(start,start+peak.peak_node+1):  # 至此，将该区间的结果存入“最终”结果中
                # 注意：“最终”list是从“起点”开始的，拟合结果直接从头开始的
                # 如：本次循环对应区间为[10,25]，则对应最终结果的[10,25],对应拟合结果的[0,15]
                self.S[i] = fitting.S[i-start]
                self.E[i] = fitting.E[i-start]
                self.I[i] = fitting.I[i-start]
                self.R[i] = fitting.R[i-start]

            if peak.exist_multipeak == 'no':  # 若不存在浪潮，退出循环
                break
            else:
                # 否则记录新浪潮节点
                print(f'新浪潮节点：{start+peak.peak_node}')
                self.new_flow_node.append(start+peak.peak_node)

                # 更新 新浪潮的起点状态，就是上次浪潮的结束状态
                S_0 = fitting.S[peak.peak_node]
                E_0 = fitting.E[peak.peak_node]
                I_0 = fitting.I[peak.peak_node]
                R_0 = fitting.R[peak.peak_node]
                start = peak.peak_node + start  # 更新起点
                continue

    def picture(self):
        t_range = np.arange(0, self.t_num)  # 时间跨度，分成一天份

        plt.plot(t_range, self.S, color='darkblue', label='Susceptible', marker='.')  # 画出易感者
        plt.plot(t_range, self.E, color='orange', label='Exposed', marker='.')  # 画出潜伏着
        plt.plot(t_range, self.I, color='red', label='Infection', marker='.')  # 画出感染者
        plt.plot(t_range, self.R, color='green', label='Recovery', marker='.')  # 画出治愈者

        plt.scatter(t_range, self.actual, color='lightblue', label='actual num')  # 画出真实感染数量

        for t in self.new_flow_node:  # 画出浪潮分割节点
            plt.plot([t, t], [0, 50000])

        plt.title('SEIR Model')
        plt.legend()
        plt.xlabel('Day')
        plt.ylabel('Number')
        plt.show()


if __name__ == '__main__':
    Epidemic_Model('疫情人数各省市数据统计列表.xlsx')

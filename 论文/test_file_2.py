import scipy.integrate as spi
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.optimize import curve_fit

t_num = 30  # 时间长度
S = [0 for i in range(t_num)]
E = [0 for i in range(t_num)]
I = [0 for i in range(t_num)]
R = [0 for i in range(t_num)]

def Fun(I, beta_i, beta_e, alpha, gamma):  # 定义拟合函数形式
    for i in range(1, 30):
        S[i] = S[i - 1] - beta_i * I[i - 1] * S[i - 1] / 150000 \
                    - beta_e * E[i - 1] * S[i - 1] / 150000

        E[i] = E[i - 1] + beta_i * I[i - 1] * S[i - 1] / 150000 \
                    + beta_e * E[i - 1] * S[i - 1] / 150000 - alpha * E[i - 1]

        I[i] = I[i - 1] + alpha * E[i - 1] - gamma * I[i - 1]

    return I


# def error(p0, S, E, I, y):  # 拟合残差
#     return np.array(Fun(p0, S, E, I)) - np.array(y)


def main():
    book = load_workbook('疫情人数各省市数据统计列表.xlsx')  # 读取数据
    sheet = book['湖北']
    t_num = 30  # 时间长度

    actual = []  # 真实的感染人数，读取数据即可
    for i in range(t_num):
        actual.append(sheet.cell(4 + i, 4).value)

    S = [0 for i in range(t_num)]
    E = [0 for i in range(t_num)]
    I = [0 for i in range(t_num)]
    R = [0 for i in range(t_num)]

    # 人群总数（武汉有5800w人口，全算上根本无法拟合，取10w看起来好很多）
    total_population = 150000

    I[0] = 41  # 感染者
    E[0] = 0  # 潜伏者
    R[0] = 0  # 恢复者
    S[0] = total_population - I[0] - E[0] - R[0]  # 易感者

    beta_i = 0.25  # 感染者传播率（即：接触数r*感染率）
    beta_e = 0.35  # 潜伏者传播率（即：接触数r*感染率）
    alpha = 0.1  # 潜伏期为10天
    gamma = 0.05  # 恢复率

    I_fit = Fun(I, beta_i, beta_e, alpha, gamma)  # 画出拟合后的曲线
    popt, pcov = curve_fit(Fun, I, I_fit)
    print(popt)


    t_range = np.arange(0, t_num)  # 时间跨度，分成一天份
    plt.plot(t_range, I_fit, label='Fitted curve')
    plt.plot(t_range, actual)
    plt.show()
    print(actual)
    print(I_fit)


if __name__ == '__main__':
    main()

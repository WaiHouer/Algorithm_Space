import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import random
import math
'''S：易感者、E：潜伏者、I：感染者、R：治愈者'''


class Metropolis_Hastings:  # Metropolis_Hastings采样算法
    # 输入：文件名，起点，终点，总人数，各群体人数初值
    def __init__(self,file_name,start,end,total_population,S_0,E_0,I_0,R_0):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['湖北']

        self.start = start  # 开始时间点
        self.end = end  # 结束时间点
        self.t_num = self.end - self.start + 1  # 时间长度

        self.actual = []  # 真实的感染人数，读取数据即可
        for i in range(self.start,self.end+1):
            self.actual.append(self.sheet.cell(4+i,4).value)

        # 人群总数（武汉有5800w人口，全算上根本无法拟合，取10w看起来好很多）
        self.total_population = total_population
        # 各群体人数初值
        self.S_0 = S_0
        self.E_0 = E_0
        self.I_0 = I_0
        self.R_0 = R_0

        self.iter = 100000  # 采样算法迭代次数

        self.para = []  # 用于存放最终拟合好的参数集合

        self.sampling()  # 采样算法主体

    def sampling(self):  # 采样算法主体
        # 参数赋初值
        beta_i = 0.35
        beta_e = 0.35
        alpha = 0.1
        gamma = 0.05

        # 初始化各群体list
        S = [0 for i in range(self.t_num)]
        E = [0 for i in range(self.t_num)]
        I = [0 for i in range(self.t_num)]
        R = [0 for i in range(self.t_num)]
        S[0] = self.S_0
        E[0] = self.E_0
        I[0] = self.I_0
        R[0] = self.R_0

        # （1）计算初始的SSE值
        SSE = self.algorithm(S,E,I,R,beta_i,beta_e,alpha,gamma)

        # （2）迭代采样
        for i in range(self.iter):
            if i % 20000 == 0:
                print(f'完成迭代{i}次')

            # （2-1）记录上一次状态
            SSE_bef = SSE  # 记录上一次SSE结果

            beta_i_bef = beta_i  # 记录上一次参数结果
            beta_e_bef = beta_e
            alpha_bef = alpha
            gamma_bef = gamma

            # （2-2）产生新解，即：从均匀分布中随机抽取，并计算新的SSE
            beta_i = random.uniform(0,1)
            beta_e = random.uniform(0,1)
            alpha = random.uniform(0,0.3)
            gamma = random.uniform(0,0.2)

            SSE = self.algorithm(S,E,I,R,beta_i,beta_e,alpha,gamma)  # 新SSE

            # （2-3）计算LR值，以一定概率接受SSE更小的结果
            LR = math.exp(min(700,(-SSE + SSE_bef)))  # 700是因为限制一下math.exp上限，防止报错
            r = random.uniform(0,1)
            if LR >= r:
                continue
            else:
                beta_i = beta_i_bef  # 若接受，更新参数结果，更新SSE结果
                beta_e = beta_e_bef
                alpha = alpha_bef
                gamma = gamma_bef
                SSE = SSE_bef

        #  （3）存入最终拟合好的参数
        self.para = [beta_i, beta_e, alpha, gamma]
        print(f'beta_i:{beta_i} ; beta_e:{beta_e} ; alpha:{alpha} ; gamma:{gamma}')
        print(f'SSE:{SSE}')

    def algorithm(self,S,E,I,R,beta_i,beta_e,alpha,gamma):  # 用于计算并返回SSE
        for i in range(1,self.t_num):
            S[i] = S[i-1] - beta_i * I[i-1] * S[i-1] / self.total_population \
                        - beta_e * E[i-1] * S[i-1] / self.total_population

            E[i] = E[i-1] + beta_i * I[i-1] * S[i-1] / self.total_population \
                        + beta_e * E[i-1] * S[i-1] / self.total_population - alpha * E[i-1]

            I[i] = I[i-1] + alpha * E[i-1] - gamma * I[i-1]

            R[i] = R[i-1] + gamma * I[i-1]

        SSE = 0
        for i in range(self.t_num):
            SSE += (I[i] - self.actual[i]) ** 2  # 这里采取的SSE就是“感染人数的均方误差”

        return SSE  # 返回SSE


if __name__ == '__main__':
    Metropolis_Hastings('疫情人数各省市数据统计列表.xlsx',0,59,150000,150000-41,0,41,0)

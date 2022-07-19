"""
画出实际传播曲线，训练集1和2的划分图
"""
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


class Actual_Trend_Show:  # 完整传染病模型
    def __init__(self,file_name):
        self.file_name = file_name  # 文件名
        self.book = load_workbook(file_name)  # 加载文件

        self.region_num = 5  # 地区数量

        self.sheet = []  # 加载每个地区（按顺序：0，1，2......）
        self.region_num = self.region_num
        for i in range(self.region_num):
            self.sheet.append(self.book[f'{i}'])  # 将每一个sheet存入列表

        self.region_name = ['' for i in range(self.region_num)]  # 记录每个州的名字，便于画图
        for i in range(self.region_num):
            self.region_name[i] = self.sheet[i].cell(1,1).value

        # 记录完整的拟合区间（如：4月13日起，前self.end - self.start + 1天）
        self.start = 0  # 开始时间点
        self.end = 319  # 结束时间点（20.4.1-21.1.15，此处为289，文件起点77）（20.4.1-21.9.1，此处为153，文件起点77）
        self.t_num = self.end - self.start + 1  # 时间长度

        self.actual = [[] for i in range(self.region_num)]  # 真实感染人数（从4月13号开始）
        for i in range(self.region_num):
            for j in range(self.t_num):
                self.actual[i].append(self.sheet[i].cell(1, self.start + j + 77).value)
        # print(self.actual)
        self.picture()

    def picture(self):
        t_range = np.arange(0, self.t_num)
        plt.plot(t_range, self.actual[0], label=f'{self.region_name[0]}', linestyle='--')
        plt.plot(t_range, self.actual[1], label=f'{self.region_name[1]}', linestyle='-')
        plt.plot(t_range, self.actual[2], label=f'{self.region_name[2]}', marker='.', markersize=5)
        plt.plot(t_range, self.actual[3], label=f'{self.region_name[3]}', linestyle='-.')
        plt.plot(t_range, self.actual[4], label=f'{self.region_name[4]}', linestyle=':')
        plt.legend(fontsize=15, facecolor='lightyellow')

        month_num = [0, 30, 61, 91, 122, 153, 183, 214, 244, 275, 306, 334]  # 画出年月日坐标（用于21-1-15训练集）
        month = ['4/1/20', '5/1/20', '6/1/20', '7/1/20', '8/1/20', '9/1/20', '10/1/20', '11/1/20', '12/1/20',
                 '1/1/21', '2/1/21', '3/1/21']
        plt.xticks(month_num, month, fontsize=15)
        plt.ylim(0,630000)  # 设置纵轴上下限
        plt.yticks(fontsize=15)

        plt.axvline(x=289, color='seagreen')  # 画出训练集分界线（用于21-1-15训练集）
        plt.axvline(x=153, color='seagreen')  # 画出训练集分界线（用于20-9-1训练集）
        plt.axvline(x=12, color='seagreen')
        plt.axvline(x=183, color='lightblue')
        plt.axvline(x=319, color='lightblue')

        # 拟合起点竖线
        # plt.annotate(f'Fitting starting point\n(12 Apr 2020)', xy=(12,130000), xytext=(20,130000),
        #              arrowprops=dict(facecolor='black', shrink=0.5, width=1, headwidth=3), fontsize=13)

        # case 1 拟合区间横线+说明
        plt.annotate(f'Fitting Period (Case 1)\n(12 Apr 2020 ~ 1 Sep 2020)', xy=(82.5, 180000), xytext=(40, 200000),
                     arrowprops=dict(facecolor='black', shrink=0.5, width=1, headwidth=1), fontsize=13)
        plt.plot([12,153], [180000, 180000], linestyle='--')
        # case 1 预测区间横线+说明
        plt.annotate(f'Prediction Period (Case 1)\n(2 Sep 2020 ~ 1 Oct 2020)', xy=(168.5, 250000), xytext=(133, 270000),
                     arrowprops=dict(facecolor='black', shrink=0.5, width=1, headwidth=1), fontsize=13)
        plt.plot([154, 183], [250000, 250000], linestyle='--')

        # case 2 拟合区间横线+说明
        plt.annotate(f'Fitting Period (Case 2)\n(12 Apr 2020 ~ 15 Jan 2021)', xy=(236, 350000), xytext=(190, 370000),
                     arrowprops=dict(facecolor='black', shrink=0.5, width=1, headwidth=1), fontsize=13)
        plt.plot([12, 289], [350000, 350000], linestyle='--')
        # case 2 预测区间横线+说明
        plt.annotate(f'Prediction Period (Case 2)\n(16 Jan 2021 ~ 15 Feb 2021)', xy=(304, 550000), xytext=(233, 570000),
                     arrowprops=dict(facecolor='black', shrink=0.5, width=1, headwidth=1), fontsize=13)
        plt.plot([289, 319], [550000, 550000], linestyle='--')

        plt.title(f'The Epidemic Development Trend', fontsize=20)
        plt.xlabel('Time point (Day)', fontsize=20)
        plt.ylabel('Actual Infected Numbers', fontsize=20)
        plt.show()


if __name__ == '__main__':
    Actual_Trend_Show('American_data.xlsx')

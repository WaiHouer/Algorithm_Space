"""
完整的传染病模型入口，包括：模型，参数拟合，多峰判断
"""
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import time
import math
from Metropolis_Hastings_Final import Metropolis_Hastings
from SEAQURD import SEAQURD
from Multipeak_judge import Multipeak_judge


class Epidemic_Model:  # 完整传染病模型
    def __init__(self,file_name):
        self.start_time = time.time()  # 计时
        self.file_name = file_name  # 文件名
        self.book = load_workbook(file_name)  # 加载文件

        self.region_num = 9  # 地区数量

        self.sheet = []  # 加载每个地区（按顺序：0，1，2......）
        self.region_num = self.region_num
        for i in range(self.region_num):
            self.sheet.append(self.book[f'{i}'])  # 将每一个sheet存入列表

        self.region_name = ['' for i in range(self.region_num)]  # 记录每个州的名字，便于画图
        for i in range(self.region_num):
            self.region_name[i] = self.sheet[i].cell(1,1).value

        # 记录完整的拟合区间（如：4月13日起，前self.end - self.start + 1天）
        self.start = 0  # 开始时间点
        self.end = 50  # 结束时间点（20.4.12-21.1.15，此处为278，文件起点88）（20.4.12-21.9.1，此处为142，文件起点88）
        self.t_num = self.end - self.start + 1  # 时间长度

        self.fitting_num = 14  # 拟合小周期

        self.predict_num = 28  # 预测未来天数

        self.actual = [[] for i in range(self.region_num)]  # 真实感染人数（从4月13号开始）
        for i in range(self.region_num):
            for j in range(self.t_num + self.predict_num):
                self.actual[i].append(self.sheet[i].cell(1, self.start + j + 88).value)
        # print(self.actual)

        # 初始化群体，用于记录“最终”的拟合结果（即：整合了多峰后的最终结果）
        self.S = np.zeros((self.region_num, self.t_num))
        self.E = np.zeros((self.region_num, self.t_num))
        self.A = np.zeros((self.region_num, self.t_num))
        self.Q = np.zeros((self.region_num, self.t_num))
        self.U = np.zeros((self.region_num, self.t_num))
        self.R = np.zeros((self.region_num, self.t_num))
        self.D = np.zeros((self.region_num, self.t_num))

        self.I = np.zeros((self.region_num, self.t_num))

        # 总人数（不变）
        self.total_population = []
        for i in range(self.region_num):
            self.total_population.append(self.sheet[i].cell(1,2).value)
        # print(self.total_population)

        self.new_flow_node = []  # 存放新浪潮开始节点（方便画图）

        print('读取数据完毕')

        self.observe = [0,'no']  # 设置一个观察收敛性的开关，看一次就行了，省的多峰判断，每判断一次看一次

        self.final_para = []  # 用于存放终末状态的拟合参数（用于预测）
        # 初始化预测阶段群体，用于记录“最终”的预测结果（针对7个不同滚动天数）
        self.S_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.E_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.A_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.Q_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.U_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.R_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.D_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]
        self.I_pre = [np.zeros((self.region_num, self.predict_num)) for i in range(7)]

        self.S_pre_direct = np.zeros((self.region_num, self.predict_num))  # 直接预测的结果（不需要的时候，视情况注释掉）
        self.E_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.A_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.Q_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.U_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.R_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.D_pre_direct = np.zeros((self.region_num, self.predict_num))
        self.I_pre_direct = np.zeros((self.region_num, self.predict_num))

        self.rolling_num = 1  # 滚动预测周期

        self.MAPE_rolling_pre = [0 for i in range(7)]  # 两种预测方式的平均绝对百分比误差（总）
        self.MAPE_direct_pre = 0
        self.MAPE_rolling_pre_list = [[0 for i in range(self.region_num)] for i in range(7)]  # 两种预测方式的平均绝对百分比误差（各区域）
        self.MAPE_direct_pre_list = [0 for i in range(self.region_num)]

        self.MAPE_fitting = 0  # 拟合的平均绝对百分比误差（总）
        self.MAPE_fitting_list = [0 for i in range(self.region_num)]  # 拟合的平均绝对百分比误差（各区域）

        self.fit_time_start = time.time()  # 对拟合时间计时
        print(f'拟合算法开始')
        self.algorithm()  # 拟合总算法
        self.mean_absolute_percentage_error_fit()  # 拟合部分MAPE计算
        print(f'拟合误差统计：')
        print(f'总体拟合MAPE：{self.MAPE_fitting}')
        print(f'各区域拟合MAPE：{self.MAPE_fitting_list}')
        self.fit_time_end = time.time()

        print(f'直接预测算法开始')
        self.direct_predict()  # 直接预测算法（一定要在滚动预测前进行，因为滚动预测会更新掉final_para）
        '---提前看直接预测的MAPE，若此处启用，则滚动后的直接预测MAPE无效，需要注意'
        self.mean_absolute_percentage_error()  # 计算MAPE
        print(f'各区域直接预测MPAE：{self.MAPE_direct_pre_list}')

        # print(f'直接预测群体数量：')  # 为了资源分配，所以显示出来，平时不用
        # print(f'S:{self.S_pre_direct}')
        # print(f'E:{self.E_pre_direct}')
        # print(f'A:{self.A_pre_direct}')
        # print(f'Q:{self.Q_pre_direct}')
        # print(f'U:{self.U_pre_direct}')
        # print(f'R:{self.R_pre_direct}')
        # print(f'D:{self.D_pre_direct}')
        print(f'拟合倒数第二个时期：（用于算U_new）')
        print(f'S:{self.S[:, -2]}')
        print(f'E:{self.E[:, -2]}')
        print(f'A:{self.A[:, -2]}')
        print(f'Q:{self.Q[:, -2]}')
        print(f'U:{self.U[:, -2]}')
        print(f'R:{self.R[:, -2]}')
        print(f'D:{self.D[:, -2]}')

        self.pre_time_start = time.time()  # 对滚动预测时间计时
        self.final_para_fixed = self.final_para  # 固定最终参数（方便针对不同滚动情况，进行初始化）
        print(f'最终参数：')
        print(self.final_para_fixed)
        self.try_roll = [ 6]  # 对滚动天数为3，7各尝试一下
        for i in self.try_roll:
            print(f'滚动预测算法开始-滚动{i+1}天')
            self.rolling_num = i + 1
            self.rolling_predict()  # 滚动预测算法

            self.mean_absolute_percentage_error()  # 计算MAPE
            print(f'滚动天数：{i + 1}天')
            print(f'总体滚动预测MAPE：{self.MAPE_rolling_pre[i]}，总体直接预测MAPE：{self.MAPE_direct_pre}')
            print(f'各区域滚动预测MPAE：{self.MAPE_rolling_pre_list[i]}')
            print(f'各区域直接预测MPAE：{self.MAPE_direct_pre_list}')
            print(self.MAPE_rolling_pre_list)
            self.final_para = self.final_para_fixed  # 重置最终参数
            self.MAPE_direct_pre = 0
            self.MAPE_direct_pre_list = [0 for i in range(self.region_num)]
        self.pre_time_end = time.time()

        self.picture()  # 画图

        self.end_time = time.time()
        print(f'SEYIAQURD-传染病模型总体运行时间为：{self.end_time - self.start_time}s')
        print(f'其中，拟合总时间为：{self.fit_time_end - self.fit_time_start}s，'
              f'滚动预测总时间为：{self.pre_time_end - self.pre_time_start}s')

    def algorithm(self):  # 拟合总算法
        # 初始化各群体数量
        E_0 = [0 for i in range(self.region_num)]
        A_0 = [0 for i in range(self.region_num)]
        Q_0 = [0 for i in range(self.region_num)]
        U_0 = [self.actual[i][0] for i in range(self.region_num)]
        R_0 = [0 for i in range(self.region_num)]
        D_0 = [0 for i in range(self.region_num)]

        S_0 = []
        for i in range(self.region_num):
            S_0.append(self.total_population[i] - U_0[i])

        s = self.start  # 初始化拟合起点（随迭代向后推动变化）
        e = self.end  # 初始化拟合终点（始终不变）
        start_list = []
        end_list = []
        while e - s + 1 >= self.fitting_num:  # 每30天为一个拟合周期，这样拟合更精准且迭代次数更少
            # 对于9月1：不足2倍周期的看做一个周期 // 对于1月15：不足1个周期的看做一个周期
            start_list.append(s)
            end_list.append(s + self.fitting_num - 1)
            s += self.fitting_num - 1
        if s != e:  # 若有剩余，代表最后不足周期，则看做最后一个周期
            start_list.append(s)
            end_list.append(e)

        for st in range(len(start_list)):  # 对每个周期进行拟合
            start = start_list[st]
            end = end_list[st]
            print(f'开始拟合，时间段：第{start + 1} - {end + 1}天')

            self.observe[0] += 1  # 观察拟合参数的周期计数+1

            while True:  # 循环：不断迭代，对新浪潮进行拟合
                # （1）MCMC算法，参数拟合
                # 输入：地区数量，文件名，起点，终点，总人数，各群体初值
                # 拟合后得到：para，为参数的集合列表
                sample = Metropolis_Hastings(self.region_num,self.book,start,end,self.total_population
                                             ,S_0,E_0,A_0,Q_0,U_0,R_0,D_0)
                # print(sample.para)

                # if self.observe[0] == 2 and self.observe[1] == 'no':  # 观察一次（只观察第n段的拟合情况）
                #     # print(sample.observe_para)
                #     self.observe_para(sample.observe_para)
                #     self.observe[1] = 'yes'

                # （2）将拟合得到的参数，输入传染病模型，从而得到各群体的拟合数量
                # 输入：地区数量，文件名，起点，终点，总人数，各群体初值，拟合好的参数
                # 模型运算后得到：各群体拟合数量
                # 注：拟合数量的list长度，随着起点的向后推移而逐渐变短
                fitting = SEAQURD(self.region_num,self.book,start,end,self.total_population
                                  ,S_0,E_0,A_0,Q_0,U_0,R_0,D_0,sample.para)

                # 将真实感染人数list进行切片，目的是与拟合list长度和对应区间保持一致
                # 同样也是随着起点向后变短
                act = []
                for i in range(len(self.actual)):
                    act.append(self.actual[i][start:end+1])
                total_I = []
                total_S = []
                total_actual = []
                for i in range(end-start+1):
                    ii = 0
                    ss = 0
                    aa = 0
                    for j in range(self.region_num):
                        ii += fitting.I[j][i]
                        ss += fitting.S[j][i]
                        aa += act[j][i]
                    total_I.append(ii)
                    total_S.append(ss)
                    total_actual.append(aa)
                # print(len(total_I),len(total_actual))

                # （3）多峰判断，确定变化点
                # 输入：真实感染人数list（切片后），该区间的拟合好的S数量，该区间的拟合好的I数量
                # 得到：新浪潮开始节点peak_node

                print('多峰算法开始')
                # peak = Multipeak_judge(total_actual,total_S,total_I)  # 总体判断多峰
                '------------------------------'
                peak = Multipeak_judge(total_actual,total_S,total_I)  # 首先，总体判断一次
                # print(f'总体判断节点：{peak.peak_node}')
                for j in range(self.region_num):
                    peak_tem = Multipeak_judge(act[j],fitting.S[j],fitting.I[j])  # 各区域分别判断，取最小值
                    # print(f'区域{j}的判断节点：{peak_tem.peak_node}')
                    if peak_tem.peak_node < peak.peak_node:
                        peak = peak_tem
                print(f'最终节点：{peak.peak_node}')

                for i in range(start,start+peak.peak_node+1):  # 至此，将该区间的结果存入“最终”结果中（修改：+1）
                    # 注意：“最终”list是从“起点”开始的，拟合结果直接从头开始的
                    # 如：本次循环对应区间为[10,25]，则对应最终结果的[10,25],对应拟合结果的[0,15]
                    for j in range(self.region_num):
                        self.S[j][i] = fitting.S[j][i-start]
                        self.E[j][i] = fitting.E[j][i-start]
                        self.A[j][i] = fitting.A[j][i-start]
                        self.Q[j][i] = fitting.Q[j][i-start]
                        self.U[j][i] = fitting.U[j][i-start]
                        self.R[j][i] = fitting.R[j][i-start]
                        self.D[j][i] = fitting.D[j][i-start]

                if peak.exist_multipeak == 'no' or start + peak.peak_node == end:  # 若不存在浪潮，退出循环
                    self.final_para = sample.para  # 记录一下最后阶段的拟合参数，用于预测未来
                    break
                else:
                    # 否则记录新浪潮节点
                    print(f'新浪潮节点：{start+peak.peak_node}')
                    self.new_flow_node.append(start+peak.peak_node)

                    # 更新 新浪潮的起点状态，就是上次浪潮的结束状态
                    for i in range(self.region_num):
                        S_0[i] = fitting.S[i][peak.peak_node]
                        E_0[i] = fitting.E[i][peak.peak_node]
                        A_0[i] = fitting.A[i][peak.peak_node]
                        Q_0[i] = fitting.Q[i][peak.peak_node]
                        U_0[i] = fitting.U[i][peak.peak_node]
                        R_0[i] = fitting.R[i][peak.peak_node]
                        D_0[i] = fitting.D[i][peak.peak_node]

                    start = peak.peak_node + start  # 更新起点
                    continue

            for i in range(self.region_num):  # 更新下一个周期的起始状态，即本周期的结束状态
                S_0[i] = self.S[i][end]
                E_0[i] = self.E[i][end]
                A_0[i] = self.A[i][end]
                Q_0[i] = self.Q[i][end]
                U_0[i] = self.U[i][end]
                R_0[i] = self.R[i][end]
                D_0[i] = self.D[i][end]
            print('beta_e', self.final_para[0])
            print('beta_a', self.final_para[1])
            print('beta_u', self.final_para[2])
            print('alpha', self.final_para[3])
            print('delta_a', self.final_para[4])
            print('delta_q', self.final_para[5])
            print('delta_u', self.final_para[6])
            print('gamma_a', self.final_para[7])
            print('gamma_q', self.final_para[8])
            print('gamma_u', self.final_para[9])
            print('p', self.final_para[10])
            print('q', self.final_para[11])
            print('c_0', self.final_para[12])

        self.I = self.A + self.Q + self.U

    def picture(self):
        t_range = np.arange(0, self.t_num)  # 时间跨度，分成一天份（用于拟合数据）
        t_pre_range = np.arange(self.end + 1, self.end + 1 + self.predict_num)  # 时间跨度（用于预测数据）

        # for t in self.new_flow_node:  # 画出浪潮分割节点
        #     plt.plot([t, t], [0, 100000])

        # 计算出各区域加总在一起的总S、E、A、Q、U、R、D、感染人数、真实人数，并画出图像
        total_S, total_E, total_A, total_Q, total_U, total_R, total_D = [],[],[],[],[],[],[]
        total_I = []
        total_actual = []
        for i in range(self.t_num):
            s, e, a, q, u, r, d = 0, 0, 0, 0, 0, 0, 0
            ii = 0
            for j in range(self.region_num):
                s += self.S[j][i]
                e += self.E[j][i]
                a += self.A[j][i]
                q += self.Q[j][i]
                u += self.U[j][i]
                r += self.R[j][i]
                d += self.D[j][i]
                ii += self.I[j][i]
            total_S.append(s), total_E.append(e), total_A.append(a), total_Q.append(q)
            total_U.append(u), total_R.append(r), total_D.append(d)
            total_I.append(ii)

        t_actual_range = np.arange(0, self.t_num + self.predict_num)  # 时间跨度（用于真实数据）
        for i in range(self.t_num + self.predict_num):  # 真实人数
            aa = 0
            for j in range(self.region_num):
                aa += self.actual[j][i]
            total_actual.append(aa)
        # plt.plot(t_range, total_S, label='Total_S', marker='.', color='orange')  # 群体过大
        # plt.plot(t_range, total_E, label='Total_E', color='red')  # 群体过大
        # plt.plot(t_range, total_A, label='Total_A', color='purple')
        # plt.plot(t_range, total_Q, label='Total_Q', color='olivedrab')
        # plt.plot(t_range, total_U, label='Total_U', color='green')
        # plt.plot(t_range, total_R, label='Total_R', color='darkblue')
        # plt.plot(t_range, total_D, label='Total_D', color='black')
        # plt.plot(t_range, total_I, label='Total_I', color='darkred')
        if self.region_num > 1:  # 画出各区域统计数据
            for i in range(self.region_num):
                # 拟合部分
                plt.plot(t_range, self.I[i],
                         label=f'{self.region_name[i]}_I')
                plt.plot(t_actual_range, self.actual[i],
                         label=f'{self.region_name[i]}_Actual', marker='.')
                # 预测部分
                for j in self.try_roll:  # 7种滚动天数情况
                    plt.plot(t_pre_range, self.I_pre[j][i],
                             label=f'{self.region_name[i]}_Rolling({j + 1} days)_Predict_I')  # 画出滚动预测结果

                plt.plot(t_pre_range, self.I_pre_direct[i], color='salmon',
                         label=f'{self.region_name[i]}_Direct_Predict_I')  # 画出直接预测结果

        # plt.plot(t_actual_range, total_actual, label='Total_Actual', color='orange', marker='.')

        # 预测总图线部分
        total_I_pre = [[] for i in range(7)]
        total_I_pre_direct = []  # 直接预测部分（其余线略，目前只统计I）
        for i in range(self.predict_num):
            ii_direct = 0
            for j in range(self.region_num):
                ii_direct += self.I_pre_direct[j][i]
            total_I_pre_direct.append(ii_direct)
        for k in self.try_roll:
            for i in range(self.predict_num):
                ii = 0
                for j in range(self.region_num):
                    ii += self.I_pre[k][j][i]
                total_I_pre[k].append(ii)

        # for k in self.try_roll:
        #     plt.plot(t_pre_range, total_I_pre[k], label=f'Rolling({k + 1} days)_Predict_Total_I')  # 画出滚动预测结果

        # plt.plot(t_pre_range, total_I_pre_direct, color='salmon', label=f'Direct_Predict_Total_I')  # 画出直接预测结果

        # month_num = [0, 19, 50, 80, 111, 142, 172, 203, 233, 264, 295]  # 画出年月日坐标（用于21-1-15训练集）
        # month = ['4/12/20', '5/1/20', '6/1/20', '7/1/20', '8/1/20', '9/1/20', '10/1/20', '11/1/20', '12/1/20',
        #          '1/1/21', '2/1/21']

        month_num = [0, 19, 50, 80, 111, 142, 172]  # 画出年月日坐标（用于20-9-1训练集）
        month = ['4/12/20', '5/1/20', '6/1/20', '7/1/20', '8/1/20', '9/1/20', '10/1/20']

        plt.xticks(month_num, month, fontsize=15)
        plt.yticks(fontsize=15)  # 设置纵轴字体大小
        # plt.axvline(x=278, color='seagreen')  # 画出训练集分界线（用于21-1-15训练集）
        # plt.axvline(x=142, color='seagreen')  # 画出训练集分界线（用于20-9-1训练集）

        # plt.title('Multipeak SEIYAQURD Model')
        plt.title(f'{self.region_name[0]} - Multipeak SEIYAQURD Model', fontsize=20)  # 单区域时，作画标题用这个

        # plt.legend(fontsize=15, facecolor='lightyellow')  # 图例显示

        plt.xlabel('Time point (Day)', fontsize=20)
        plt.ylabel('Infected Numbers', fontsize=20)
        plt.show()

    def observe_para(self,observe_para):  # 函数：用于显示参数变化图
        x = np.arange(0, len(observe_para[0][0]))
        for i in range(self.region_num):
            plt.plot(x, observe_para[i][0], label='beta_e')
            plt.plot(x, observe_para[i][1], label='beta_a')
            plt.plot(x, observe_para[i][2], label='beta_u')
            plt.plot(x, observe_para[i][3], label='alpha')
            plt.plot(x, observe_para[i][4], label='delta_a')
            plt.plot(x, observe_para[i][5], label='delta_q')
            plt.plot(x, observe_para[i][6], label='delta_u')
            plt.plot(x, observe_para[i][7], label='gamma_a')
            plt.plot(x, observe_para[i][8], label='gamma_q')
            plt.plot(x, observe_para[i][9], label='gamma_u')
            plt.title(f'Observe Parameters - {self.region_name[i]}')
            plt.legend(fontsize=10, facecolor='lightyellow')
            plt.legend()
            plt.xlabel('Iteration')
            plt.ylabel('Parameters')
            plt.show()

    def rolling_predict(self):
        S_0 = [0 for i in range(self.region_num)]  # 滚动拟合起点（每次前进 self.rolling_num 天，持续15天）
        E_0 = [0 for i in range(self.region_num)]
        A_0 = [0 for i in range(self.region_num)]
        Q_0 = [0 for i in range(self.region_num)]
        U_0 = [0 for i in range(self.region_num)]
        R_0 = [0 for i in range(self.region_num)]
        D_0 = [0 for i in range(self.region_num)]

        S_0_pre = [0 for i in range(self.region_num)]  # 滚动预测起点（为每次拟合的最后一天）
        E_0_pre = [0 for i in range(self.region_num)]
        A_0_pre = [0 for i in range(self.region_num)]
        Q_0_pre = [0 for i in range(self.region_num)]
        U_0_pre = [0 for i in range(self.region_num)]
        R_0_pre = [0 for i in range(self.region_num)]
        D_0_pre = [0 for i in range(self.region_num)]

        start = self.end + 1 - self.fitting_num  # 拟合起点（从结束点前15天开始）
        end = self.end

        rolling_num = self.rolling_num

        iteration = math.ceil(self.predict_num / self.rolling_num)  # 需要滚动的次数=预测天数/滚动天数，向上取整

        for i in range(iteration):
            if i == iteration - 1:  # 最后一个滚动周期（不足的算作一次）
                rolling_num = self.predict_num - i * rolling_num
            print(f'开始预测第{i * self.rolling_num + 1} ~ {i * self.rolling_num + 1 + rolling_num - 1}天')
            if i == 0:
                for j in range(self.region_num):
                    S_0_pre[j] = self.S[j][-1]  # 拟合的最后一天为预测起点
                    E_0_pre[j] = self.E[j][-1]
                    A_0_pre[j] = self.A[j][-1]
                    Q_0_pre[j] = self.Q[j][-1]
                    U_0_pre[j] = self.U[j][-1]
                    R_0_pre[j] = self.R[j][-1]
                    D_0_pre[j] = self.D[j][-1]

                fitting = SEAQURD(self.region_num,self.book,end,end + rolling_num,self.total_population,
                                  S_0_pre,E_0_pre,A_0_pre,Q_0_pre,U_0_pre,R_0_pre,D_0_pre,self.final_para)
                for j in range(self.region_num):  # 存入最终预测结果
                    for b in range(rolling_num):
                        self.S_pre[self.rolling_num - 1][j][i + b] = fitting.S[j][1 + b]  # 除去第一天是衔接天
                        self.E_pre[self.rolling_num - 1][j][i + b] = fitting.E[j][1 + b]
                        self.A_pre[self.rolling_num - 1][j][i + b] = fitting.A[j][1 + b]
                        self.Q_pre[self.rolling_num - 1][j][i + b] = fitting.Q[j][1 + b]
                        self.U_pre[self.rolling_num - 1][j][i + b] = fitting.U[j][1 + b]
                        self.R_pre[self.rolling_num - 1][j][i + b] = fitting.R[j][1 + b]
                        self.D_pre[self.rolling_num - 1][j][i + b] = fitting.D[j][1 + b]

                start += rolling_num  # 滚动
                end += rolling_num

                for j in range(self.region_num):
                    S_0[j] = self.S[j][start]  # 更新滚动拟合起点（向后推rolling_num天）
                    E_0[j] = self.E[j][start]
                    A_0[j] = self.A[j][start]
                    Q_0[j] = self.Q[j][start]
                    U_0[j] = self.U[j][start]
                    R_0[j] = self.R[j][start]
                    D_0[j] = self.D[j][start]
            else:
                start_train = start
                end_train = end

                S_pre_tem = np.zeros((self.region_num, self.fitting_num))  # 临时记录器（用于记录本次拟合结果）
                E_pre_tem = np.zeros((self.region_num, self.fitting_num))  # 其中，第 rolling_num 个点用于更新拟合起点
                A_pre_tem = np.zeros((self.region_num, self.fitting_num))  # 其中，最后-1个点用于更新预测起点
                Q_pre_tem = np.zeros((self.region_num, self.fitting_num))
                U_pre_tem = np.zeros((self.region_num, self.fitting_num))
                R_pre_tem = np.zeros((self.region_num, self.fitting_num))
                D_pre_tem = np.zeros((self.region_num, self.fitting_num))

                while True:
                    sample = Metropolis_Hastings(self.region_num,self.book,start_train,end_train,
                                                 self.total_population,
                                                 S_0,E_0,A_0,Q_0,U_0,R_0,D_0)
                    fitting = SEAQURD(self.region_num,self.book,start_train,end_train,self.total_population,
                                      S_0,E_0,A_0,Q_0,U_0,R_0,D_0,sample.para)

                    act = []
                    for a in range(len(self.actual)):
                        act.append(self.actual[a][start_train:end_train + 1])
                    total_I = []
                    total_S = []
                    total_actual = []
                    for a in range(end_train - start_train + 1):
                        ii = 0
                        ss = 0
                        aa = 0
                        for b in range(self.region_num):
                            ii += fitting.I[b][a]
                            ss += fitting.S[b][a]
                            aa += act[b][a]
                        total_I.append(ii)
                        total_S.append(ss)
                        total_actual.append(aa)

                    print('多峰算法开始')
                    # peak = Multipeak_judge(total_actual,total_S,total_I)
                    '------------------------------'
                    peak = Multipeak_judge(total_actual, total_S, total_I)  # 首先，总体判断一次
                    # print(f'总体判断节点：{peak.peak_node}')
                    for j in range(self.region_num):
                        peak_tem = Multipeak_judge(act[j], fitting.S[j], fitting.I[j])  # 各区域分别判断，取最小值
                        # print(f'区域{j}的判断节点：{peak_tem.peak_node}')
                        if peak_tem.peak_node < peak.peak_node:
                            peak = peak_tem
                    print(f'最终节点：{peak.peak_node}')

                    for a in range(start_train,start_train + peak.peak_node + 1):  # 存入临时记录器中
                        for b in range(self.region_num):
                            S_pre_tem[b][a - start] = fitting.S[b][a - start_train]
                            E_pre_tem[b][a - start] = fitting.E[b][a - start_train]
                            A_pre_tem[b][a - start] = fitting.A[b][a - start_train]
                            Q_pre_tem[b][a - start] = fitting.Q[b][a - start_train]
                            U_pre_tem[b][a - start] = fitting.U[b][a - start_train]
                            R_pre_tem[b][a - start] = fitting.R[b][a - start_train]
                            D_pre_tem[b][a - start] = fitting.D[b][a - start_train]

                    if peak.exist_multipeak == 'no' or start_train + peak.peak_node == end_train:
                        self.final_para = sample.para
                        break
                    else:
                        # print(f'新浪潮节点：{start_train+peak.peak_node}')
                        for a in range(self.region_num):
                            S_0[a] = fitting.S[a][peak.peak_node]
                            E_0[a] = fitting.E[a][peak.peak_node]
                            A_0[a] = fitting.A[a][peak.peak_node]
                            Q_0[a] = fitting.Q[a][peak.peak_node]
                            U_0[a] = fitting.U[a][peak.peak_node]
                            R_0[a] = fitting.R[a][peak.peak_node]
                            D_0[a] = fitting.D[a][peak.peak_node]
                        start_train = peak.peak_node + start_train
                        continue

                for a in range(self.region_num):  # 更新预测起点（即：本次拟合最后一个点）
                    S_0_pre[a] = S_pre_tem[a][-1]
                    E_0_pre[a] = E_pre_tem[a][-1]
                    A_0_pre[a] = A_pre_tem[a][-1]
                    Q_0_pre[a] = Q_pre_tem[a][-1]
                    U_0_pre[a] = U_pre_tem[a][-1]
                    R_0_pre[a] = R_pre_tem[a][-1]
                    D_0_pre[a] = D_pre_tem[a][-1]

                fitting = SEAQURD(self.region_num, self.book, end, end + rolling_num, self.total_population,
                                  S_0_pre, E_0_pre, A_0_pre, Q_0_pre, U_0_pre, R_0_pre, D_0_pre, self.final_para)
                for j in range(self.region_num):  # 存入最终预测结果
                    for b in range(rolling_num):
                        self.S_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.S[j][1 + b]
                        self.E_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.E[j][1 + b]
                        self.A_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.A[j][1 + b]
                        self.Q_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.Q[j][1 + b]
                        self.U_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.U[j][1 + b]
                        self.R_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.R[j][1 + b]
                        self.D_pre[self.rolling_num - 1][j][i * self.rolling_num + b] = fitting.D[j][1 + b]

                start += rolling_num  # 滚动
                end += rolling_num

                for a in range(self.region_num):  # 更新下一次拟合起点（即：本次拟合的第 rolling_num 个点）
                    S_0[a] = S_pre_tem[a][rolling_num]
                    E_0[a] = E_pre_tem[a][rolling_num]
                    A_0[a] = A_pre_tem[a][rolling_num]
                    Q_0[a] = Q_pre_tem[a][rolling_num]
                    U_0[a] = U_pre_tem[a][rolling_num]
                    R_0[a] = R_pre_tem[a][rolling_num]
                    D_0[a] = D_pre_tem[a][rolling_num]

        self.I_pre[self.rolling_num - 1] = self.A_pre[self.rolling_num - 1] + self.Q_pre[self.rolling_num - 1]\
                                           + self.U_pre[self.rolling_num - 1]  # 计算I值

    def direct_predict(self):
        S_0 = [0 for i in range(self.region_num)]
        E_0 = [0 for i in range(self.region_num)]
        A_0 = [0 for i in range(self.region_num)]
        Q_0 = [0 for i in range(self.region_num)]
        U_0 = [0 for i in range(self.region_num)]
        R_0 = [0 for i in range(self.region_num)]
        D_0 = [0 for i in range(self.region_num)]

        for i in range(self.region_num):
            S_0[i] = self.S[i][-1]
            E_0[i] = self.E[i][-1]
            A_0[i] = self.A[i][-1]
            Q_0[i] = self.Q[i][-1]
            U_0[i] = self.U[i][-1]
            R_0[i] = self.R[i][-1]
            D_0[i] = self.D[i][-1]

        print(f'拟合最后时期群体：')  # 为了资源分配显示出来，平时不用
        print(f'S:{S_0}')
        print(f'E:{E_0}')
        print(f'A:{A_0}')
        print(f'Q:{Q_0}')
        print(f'U:{U_0}')
        print(f'R:{R_0}')
        print(f'D:{D_0}')

        fitting = SEAQURD(self.region_num, self.book, self.end, self.end + self.predict_num,
                          self.total_population, S_0, E_0, A_0, Q_0, U_0, R_0, D_0, self.final_para)

        for i in range(self.region_num):
            for j in range(self.predict_num):
                self.S_pre_direct[i][j] = fitting.S[i][1 + j]  # 除去第一个点为衔接点
                self.E_pre_direct[i][j] = fitting.E[i][1 + j]
                self.A_pre_direct[i][j] = fitting.A[i][1 + j]
                self.Q_pre_direct[i][j] = fitting.Q[i][1 + j]
                self.U_pre_direct[i][j] = fitting.U[i][1 + j]
                self.R_pre_direct[i][j] = fitting.R[i][1 + j]
                self.D_pre_direct[i][j] = fitting.D[i][1 + j]

        self.I_pre_direct = self.A_pre_direct + self.Q_pre_direct + self.U_pre_direct  # 计算I值

    def mean_absolute_percentage_error(self):

        total_actual = []  # 对预测天数内的值求和（即：总实际值、总滚动预测值、总直接预测值）
        total_rolling = []
        total_direct = []
        for i in range(self.predict_num):
            aa, rr, dd = 0, 0, 0
            for j in range(self.region_num):
                aa += self.actual[j][self.end + 1 + i]
                rr += self.I_pre[self.rolling_num - 1][j][i]  # 只计算当前滚动情况
                dd += self.I_pre_direct[j][i]
            total_actual.append(aa)
            total_rolling.append(rr)
            total_direct.append(dd)

        for i in range(self.region_num):
            for j in range(self.predict_num):
                # 直接预测（各区域）累积误差百分比
                self.MAPE_direct_pre_list[i] += math.fabs(self.actual[i][self.end + 1 + j] - self.I_pre_direct[i][j])\
                                                / self.actual[i][self.end + 1 + j]
                # 滚动预测（各区域）累积误差百分比
                self.MAPE_rolling_pre_list[self.rolling_num - 1][i] += math.fabs(self.actual[i][self.end + 1 + j]
                                                                                 - self.I_pre[self.rolling_num - 1][i][j]) \
                                                    / self.actual[i][self.end + 1 + j]
            self.MAPE_rolling_pre_list[self.rolling_num - 1][i] /= self.predict_num  # 均值
            self.MAPE_direct_pre_list[i] /= self.predict_num  # 均值

        for i in range(self.predict_num):
            # 直接预测（总体）累积误差百分比
            self.MAPE_direct_pre += math.fabs(total_actual[i] - total_direct[i]) / total_actual[i]
            # 滚动预测（总体）累积误差百分比
            self.MAPE_rolling_pre[self.rolling_num - 1] += math.fabs(total_actual[i] - total_rolling[i]) / total_actual[i]
        self.MAPE_direct_pre /= self.predict_num  # 均值
        self.MAPE_rolling_pre[self.rolling_num - 1] /= self.predict_num  # 均值

    def mean_absolute_percentage_error_fit(self):
        total_actual = []
        total_I = []
        for i in range(self.t_num):
            aa, ii = 0, 0
            for j in range(self.region_num):
                aa += self.actual[j][i]
                ii += self.I[j][i]
            total_actual.append(aa)
            total_I.append(ii)

        for i in range(self.region_num):
            for j in range(self.t_num):
                # 拟合（各区域）累积误差百分比
                self.MAPE_fitting_list[i] += math.fabs(self.actual[i][j] - self.I[i][j]) / self.actual[i][j]
            self.MAPE_fitting_list[i] /= self.t_num

        for i in range(self.t_num):
            # 拟合（总体）累积误差百分比
            self.MAPE_fitting += math.fabs(total_actual[i] - total_I[i]) / total_actual[i]
        self.MAPE_fitting /= self.t_num


if __name__ == '__main__':
    Epidemic_Model('American_Census_Region_data.xlsx')

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
'''S：易感者、E：潜伏者、I：感染者、R：治愈者'''


class SEIR:  # SEIR传染病模型
    def __init__(self,file_name,start,end,total_population,S_0,E_0,I_0,R_0,para):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['湖北']

        self.start = start  # 开始时间点
        self.end = end  # 结束时间点
        self.t_num = self.end - self.start + 1  # 时间长度

        self.para = para  # 参数列表

        self.actual = []  # 真实的感染人数，读取数据即可
        for i in range(self.start,self.end+1):
            self.actual.append(self.sheet.cell(4+i,4).value)

        self.S = [0 for i in range(self.t_num)]
        self.E = [0 for i in range(self.t_num)]
        self.I = [0 for i in range(self.t_num)]
        self.R = [0 for i in range(self.t_num)]

        # 人群总数（武汉有5800w人口，全算上根本无法拟合，取10w看起来好很多）
        self.total_population = total_population

        self.I[0] = I_0  # 感染者
        self.E[0] = E_0  # 潜伏者
        self.R[0] = R_0  # 恢复者
        self.S[0] = S_0  # 易感者

        self.beta_i, self.beta_e, self.alpha, self.gamma = para
        # self.beta_i = 0.6166705369207879  # 感染者传播率（即：接触数r*感染率）
        # self.beta_e = 0.3779508896534194  # 潜伏者传播率（即：接触数r*感染率）
        # self.alpha = 0.044910060076917224  # 潜伏期为10天
        # self.gamma = 0.02835410545944275  # 恢复率

        self.algorithm()
        # self.picture()

    def algorithm(self):
        for i in range(1,self.t_num):
            self.S[i] = self.S[i-1] - self.beta_i * self.I[i-1] * self.S[i-1] / self.total_population \
                        - self.beta_e * self.E[i-1] * self.S[i-1] / self.total_population

            self.E[i] = self.E[i-1] + self.beta_i * self.I[i-1] * self.S[i-1] / self.total_population \
                        + self.beta_e * self.E[i-1] * self.S[i-1] / self.total_population - self.alpha * self.E[i-1]

            self.I[i] = self.I[i-1] + self.alpha * self.E[i-1] - self.gamma * self.I[i-1]

            self.R[i] = self.R[i-1] + self.gamma * self.I[i-1]

    def picture(self):
        t_range = np.arange(0, self.t_num)  # 时间跨度，分成一天份

        plt.plot(t_range, self.S, color='darkblue', label='Susceptible', marker='.')  # 画出易感者
        plt.plot(t_range, self.E, color='orange', label='Exposed', marker='.')  # 画出潜伏着
        plt.plot(t_range, self.I, color='red', label='Infection', marker='.')  # 画出感染者
        plt.plot(t_range, self.R, color='green', label='Recovery', marker='.')  # 画出治愈者

        plt.scatter(t_range, self.actual, color='lightblue', label='actual num')  # 画出真实感染数量

        plt.title('SEIR Model')
        plt.legend()
        plt.xlabel('Day')
        plt.ylabel('Number')
        plt.show()


if __name__ == '__main__':
    SEIR('疫情人数各省市数据统计列表.xlsx',0,59,150000,150000-41,0,41,0,
         [0.6166705369207879,0.3779508896534194,0.044910060076917224,0.02835410545944275])

# -*- coding: utf-8 -*-
"""
@author: Wai_Hou_er（cd）
"""
# 肘部方法
from openpyxl import load_workbook
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt


book = load_workbook('2019.xlsx')
sheet = book['Sheet2']

df_features = []  # 读入数据 '利用SSE选择k'
for i in range(3,2500):
    tem = []
    for j in range(3,15):
        if j != 5 and j != 9:
            tem.append(sheet.cell(i,j).value)

    df_features.append(tem)

SSE = []  # 存放每次结果的误差平方和
for k in range(5, 15):
    estimator = KMeans(n_clusters=k)  # 构造聚类器
    estimator.fit(df_features)
    SSE.append(estimator.inertia_)  # estimator.inertia_获取聚类准则的总和
X = range(5, 15)
plt.xlabel('k')
plt.ylabel('SSE')
plt.plot(X, SSE, 'o-')
plt.show()

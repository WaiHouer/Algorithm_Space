import numpy as np
import matplotlib.pyplot as plt
from sklearn.linear_model import BayesianRidge
from openpyxl import load_workbook


class Bayesian_Ridge_Regression:  # 贝叶斯岭回归
    def __init__(self,file_name):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['湖北']
        self.num = 30  # 历史数据数量
        self.num_future = 30  # 需要预测的未来天数

        self.x = [i for i in range(1,self.num+1)]  # 初始化训练集x，由于只包含天数，所以暂时可以直接1-30的list形式
        self.y = [0 for i in range(self.num)]      # 初始化训练集y，list形式即可
        # 同理：需要预测的测试集
        self.x_test = [i for i in range(self.num+1,self.num+1+self.num_future)]
        self.y_test = [0 for i in range(self.num_future)]

        for i in range(len(self.y)):  # 录入表格数据
            self.y[i] = self.sheet.cell(4+i,4).value

        for i in range(len(self.y_test)):  # 录入表格数据
            self.y_test[i] = self.sheet.cell(self.num_future+4+i,4).value

        self.regression()

    def regression(self):  # 回归算法主体

        n_order = 3  # 多项式最高次幂
        x_train = np.vander(self.x, n_order+1, increasing=True)  # 生成最高n_order次幂的，升序的，范德蒙矩阵
        y_train = self.y
        x_test = np.vander(self.x_test, n_order+1, increasing=True)

        # 注意：此时 x_train、x_test 的格式变成了array，列为特征（即多项式各项），行为样本
        # y_train不用管，[y,y,y,y,y,y,y,y,y]这种形式就可以

        # tol：如果w收敛，则停止算法；fit_intercept：是否计算该模型的截距；compute_score：如果为True，则计算模型每一步的目标函数
        reg = BayesianRidge(tol=1e-6, fit_intercept=False, compute_score=True)
        reg.fit(x_train, y_train)  # 用训练集训练

        y_predict_train = reg.predict(x_train)  # 对训练集拟合一下，导出y值

        y_predict_future = reg.predict(x_test)  # 对测试集拟合一下，导出y值

        # 画图
        plt.scatter(self.x+self.x_test, self.y+self.y_test, color='lightblue', label='actual num')
        plt.plot(self.x, y_predict_train, color='red', label='fit curve')
        plt.plot(self.x_test, y_predict_future, color='palegreen', label='predict curve')

        plt.legend()
        plt.show()


if __name__ == '__main__':
    Bayesian_Ridge_Regression('疫情人数各省市数据统计列表.xlsx')

"""
多元线性回归——正规方程法
"""
'''具体参考：笔记-（4）正规方程法'''
'''有两行小报错，不用管，是QQ输入法造成的'''
import matplotlib.pyplot as plt
import numpy as np
from numpy import *
from openpyxl import load_workbook
from List_0 import build_list


class Normal_Equation:
    def __init__(self,data):
        self.origin_data = data
        self.m = len(data['x0'])
        self.n = len(data) - 1

        self.data = build_list(self.m,self.n,0)
        for i in range(self.m):
            for j in range(self.n):
                self.data[i][j] = data[f'x{j}'][i]
        self.data = mat(self.data)

        self.y = mat(data['y']).T

        self.sita = []

        self.algorithm()

    def algorithm(self):
        self.sita = (self.data.T * self.data).I * self.data.T * self.y
        print(f'回归结果：{self.sita}')
        for i in range(self.m):
            print(self.sita.T * mat([self.origin_data['x0'][i], self.origin_data['x1'][i], self.origin_data['x2'][i],
                                     self.origin_data['x3'][i], self.origin_data['x4'][i]]).T,self.origin_data['y'][i])


if __name__ == '__main__':
    example = load_workbook('example_1.xlsx')
    sheet = example['1']
    x1 = []
    x2 = []
    x3 = []
    x4 = []
    y = []
    for k in range(2,119,4):
        x1 += [sheet.cell(k,1).value]
        x2 += [sheet.cell(k,2).value]
        x3 += [sheet.cell(k,3).value]
        x4 += [sheet.cell(k,4).value]
        y += [sheet.cell(k,9).value]
    exam_dict = {'x0': np.ones(30), 'x1': x1, 'x2': x2, 'x3': x3, 'x4': x4, 'y': y}

    Normal_Equation(exam_dict)

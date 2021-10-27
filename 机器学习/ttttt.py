import scipy.integrate as spi
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


class SIR:
    def __init__(self,file_name):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['湖北']
        self.t_num = 60

        self.actual = []
        for i in range(self.t_num):
            self.actual.append(self.sheet.cell(4+i,4).value)

        print(self.actual)

        self.total_population = 100000
        # β为传染率系数
        self.beta = 0.4
        # gamma为恢复率系数
        self.gamma = 0.05
        # I_0为感染者的初始人数
        self.I_0 = 41
        # R_0为治愈者的初始人数
        self.R_0 = 0
        # S_0为易感者的初始人数
        self.S_0 = self.total_population - self.I_0 - self.R_0

        self.INI = (self.S_0, self.I_0, self.R_0)

        self.algorithm()

    def function(self,initial_value,_):
        y = np.zeros(3)
        x = initial_value

        y[0] = -(self.beta * x[0] * x[1]) / self.total_population
        y[1] = (self.beta * x[0] * x[1]) / self.total_population - self.gamma * x[1]
        y[2] = self.gamma * x[1]

        return y

    def algorithm(self):
        t_range = np.arange(0, self.t_num)

        res = spi.odeint(self.function, self.INI, t_range)

        plt.plot(res[:, 0], color='darkblue', label='Susceptible', marker='.')
        plt.plot(res[:, 1], color='red', label='Infection', marker='.')
        plt.plot(res[:, 2], color='green', label='Recovery', marker='.')

        plt.scatter(t_range, self.actual, color='lightblue', label='actual num')

        plt.title('SIR Model')
        plt.legend()
        plt.xlabel('Day')
        plt.ylabel('Number')
        plt.show()


if __name__ == '__main__':
    SIR('疫情人数各省市数据统计列表.xlsx')

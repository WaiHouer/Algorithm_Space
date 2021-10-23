import numpy as np
import matplotlib.pyplot as plt
from sklearn.svm import SVR
from openpyxl import load_workbook
'''多项式核函数（Poly）的SVR：可以用'''


class Poly_SVR:  # 多项式核函数（Poly）的SVR
    def __init__(self,file_name):
        self.book = load_workbook(file_name)
        self.sheet = self.book['湖北']
        self.num = 30  # 历史数据数量
        self.num_future = 30  # 需要预测的未来天数

        self.x = [i for i in range(1,self.num+1)]  # 初始化训练集x，初步写出前30天作为特征
        self.y = []  # 初始化训练集y，list形式即可
        # 同理：需要预测的测试集
        self.x_test = [i for i in range(self.num+1,self.num+1+self.num_future)]
        self.y_test = []

        for i in range(self.num):  # 录入表格数据
            self.y.append(self.sheet.cell(4+i,4).value)

        for i in range(self.num_future):
            self.y_test.append(self.sheet.cell(self.num+4+i,4).value)

        self.regression()

    def regression(self):  # 回归算法主体

        x_train = np.array(self.x).reshape((self.num,1))
        y_train = self.y
        x_test = np.array(self.x_test).reshape((self.num_future,1))

        # 注意：此时 x_train、x_test 的格式变成了array，列为特征（即多项式各项），行为样本
        # y_train不用管，[y,y,y,y,y,y,y,y,y]这种形式就可以

        # kernel：指定内核；C：正则化系数；gamma：核系数；coef0：核函数独立项
        reg = SVR(kernel='poly', C=100, degree=3, coef0=1)
        reg.fit(x_train, y_train)  # 用训练集训练

        y_predict_train = reg.predict(x_train)  # 对训练集拟合一下，导出y值

        y_predict_future = reg.predict(x_test)  # 对测试集拟合一下，导出y值

        # 画图
        plt.scatter(self.x+self.x_test, self.y+self.y_test, color='lightblue', label='actual num')
        plt.plot(self.x, y_predict_train, color='red', label='fit curve')
        plt.plot(self.x_test, y_predict_future, color='palegreen', label='predict curve')

        plt.legend()
        plt.show()


if __name__ == '__main__':
    Poly_SVR('疫情人数各省市数据统计列表.xlsx')

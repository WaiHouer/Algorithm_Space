import numpy as np
import matplotlib.pyplot as plt
from sklearn.kernel_ridge import KernelRidge
from openpyxl import load_workbook
'''内核（多项式内核）岭回归：可以用'''


class KRR:
    def __init__(self,file_name):
        self.book = load_workbook(file_name)
        self.sheet = self.book['湖北']
        self.num = 30
        self.num_future = 30

        self.x = [i for i in range(1,self.num+1)]
        self.y = []

        self.x_test = [i for i in range(self.num+1,self.num+1+self.num_future)]
        self.y_test = []

        for i in range(self.num):
            self.y.append(self.sheet.cell(4+i,4).value)

        for i in range(self.num_future):
            self.y_test.append(self.sheet.cell(self.num+4+i,4).value)

        self.regression()

    def regression(self):
        x_train = np.array(self.x).reshape((self.num,1))
        y_train = self.y
        x_test = np.array(self.x_test).reshape((self.num_future,1))

        # kernel：指定内核；C：正则化系数；gamma：核系数；epsilon：损失值与实际值之间的距离
        reg = KernelRidge(kernel='poly')
        reg.fit(x_train, y_train)

        y_predict_train = reg.predict(x_train)

        y_predict_future = reg.predict(x_test)

        plt.scatter(self.x+self.x_test, self.y+self.y_test, color='lightblue', label='actual num')
        plt.plot(self.x, y_predict_train, color='red', label='fit curve')
        plt.plot(self.x_test, y_predict_future, color='palegreen', label='predict curve')

        plt.legend()
        plt.show()


if __name__ == '__main__':
    KRR('疫情人数各省市数据统计列表.xlsx')

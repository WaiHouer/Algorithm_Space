"""
多元线性回归——梯度下降法
"""
'''具体参考：笔记-线性回归例子-拥有多个特征的房价预测问题'''
'''有两行小报错，不用管，是QQ输入法造成的'''
'''考虑了：（1）特征缩放，（2）画图debugging'''
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as op
from openpyxl import load_workbook
from numpy import *


class Multiple_Linear_Regression:
    def __init__(self,data):  # 传入训练集数据
        self.data = data         # 训练集
        self.m = len(data['x1'])  # 训练集数据个数
        self.n = len(data) - 1    # 训练集特征个数（包括默认x0）

        self.miu = []
        self.max_min = []
        for i in range(self.n):
            self.miu.append(np.mean(self.data[f'x{i}']))
            self.max_min.append(max(self.data[f'x{i}']) - min(self.data[f'x{i}']))

        self.alpha = 0.01  # learning rate
        self.sita = np.zeros(self.n)

        self.iteration = 5000  # 循环次数

        self.algorithm()  # 主函数

    def algorithm(self):
        real_iter = 0
        for i in range(self.iteration):
            # 计算偏导
            tem = np.zeros(self.n)
            for j in range(self.n):
                tem[j] = self.sita[j] - self.alpha * self.partial_derivative(j)
            if self.sita[0] == tem[0] and self.sita[1] == tem[1] and self.sita[2] == tem[2] and self.sita[3] == tem[3] \
                    and self.sita[4] == tem[4]:
                break
            for j in range(self.n):
                self.sita[j] = tem[j]

            real_iter = i
        print(f'回归结果：{self.sita[0], self.sita[1],self.sita[2],self.sita[3],self.sita[4]}')
        print(f'实际循环次数：{real_iter}')
        for i in range(self.m):
            print(mat(self.sita) * mat([self.data['x0'][i], self.data['x1'][i], self.data['x2'][i],
                                     self.data['x3'][i], self.data['x4'][i]]).T, self.data['y'][i])

    def partial_derivative(self,para_num):
        result = 0
        # 均值归一！！！！！！！！！！！！！！
        for i in range(self.m):
            x_0 = self.data['x0'][i]
            x_1 = (self.data['x1'][i] - self.miu[1]) / self.max_min[1]
            x_2 = (self.data['x2'][i] - self.miu[2]) / self.max_min[2]
            x_3 = (self.data['x3'][i] - self.miu[3]) / self.max_min[3]
            x_4 = (self.data['x4'][i] - self.miu[4]) / self.max_min[4]
            if para_num == 0:
                xx = 1
            else:
                xx = (self.data[f'x{para_num}'][i] - self.miu[para_num]) / self.max_min[para_num]
            result += (self.sita[0] * x_0 + self.sita[1] * x_1 + self.sita[2] * x_2 + self.sita[3] * x_3
                       + self.sita[4] * x_4 - self.data['y'][i]) * xx
        result *= 1 / self.m
        return result


if __name__ == '__main__':
    example = load_workbook('example_1.xlsx')
    sheet = example['1']
    x1 = []
    x2 = []
    x3 = []
    x4 = []
    y = []
    for k in range(2,119,4):
        x1 += [sheet.cell(k,1).value]
        x2 += [sheet.cell(k,2).value]
        x3 += [sheet.cell(k,3).value]
        x4 += [sheet.cell(k,4).value]
        y += [sheet.cell(k,9).value]
    exam_dict = {'x0': np.ones(30), 'x1': x1, 'x2': x2, 'x3': x3, 'x4': x4, 'y': y}

    Multiple_Linear_Regression(exam_dict)

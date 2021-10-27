import scipy.integrate as spi
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
'''S：易感者、E：潜伏者、I：感染者、T：接受治疗者、R：治愈者'''


class SEITR:  # SEITR传染病模型
    def __init__(self,file_name):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['湖北']
        self.t_num = 60  # 时间长度

        self.actual = []  # 真实的感染人数，读取数据即可
        for i in range(self.t_num):
            self.actual.append(self.sheet.cell(4+i,4).value)

        # 人群总数（武汉有5800w人口，全算上根本无法拟合，取10w看起来好很多）
        self.total_population = 100000
        # β为传染率系数
        self.beta = 0.25
        # gamma为恢复率系数
        self.gamma = 0.04
        # delta为受到治疗系数
        self.delta = 0.1
        # Te为疾病潜伏期
        self.Te = 10
        # I_0为感染者的初始人数
        self.I_0 = 41
        # R_0为治愈者的初始人数
        self.R_0 = 0
        # E_0为潜伏者的初始人数
        self.E_0 = 0
        # T_0为治疗中的初始人数
        self.T_0 = 0
        # S_0为易感者的初始人数
        self.S_0 = self.total_population - self.I_0 - self.R_0 - self.E_0 - self.T_0

        # SIR组合起来的数组
        self.INI = (self.S_0, self.E_0, self.I_0, self.R_0, self.T_0)

        self.algorithm()

    def function(self,initial_value,_):  # 方法：微分方程表达式（暂时不懂具体，不过先用着）
        y = np.zeros(5)
        x = initial_value

        y[0] = - (self.beta * x[0] * (x[2] + x[1])) / self.total_population  # 易感个体变化
        y[1] = (self.beta * x[0] * (x[2] + x[1])) / self.total_population - x[1] / self.Te  # 潜伏个体变化
        y[2] = x[1] / self.Te - self.delta * x[2]  # 感染未住院
        y[3] = self.gamma * x[4]  # 治愈个体变化
        y[4] = self.delta * x[2] - self.gamma * x[4]  # 治疗中个体变化

        return y

    def algorithm(self):
        t_range = np.arange(0, self.t_num)  # 时间跨度，分成一天份

        res = spi.odeint(self.function, self.INI, t_range)  # 求解微分方程

        plt.plot(res[:, 0], color='darkblue', label='Susceptible', marker='.')  # 画出易感者
        plt.plot(res[:, 1], color='orange', label='Exposed', marker='.')  # 画出潜伏着
        plt.plot(res[:, 2], color='red', label='Infection', marker='.')  # 画出感染者
        plt.plot(res[:, 3], color='green', label='Recovery', marker='.')  # 画出治愈者
        plt.plot(res[:, 4], color='purple', label='Under Treatment', marker='.')  # 画出住院者

        plt.scatter(t_range, self.actual, color='lightblue', label='actual num')  # 画出真实感染数量

        plt.title('SEITR Model')
        plt.legend()
        plt.xlabel('Day')
        plt.ylabel('Number')
        plt.show()


if __name__ == '__main__':
    SEITR('疫情人数各省市数据统计列表.xlsx')

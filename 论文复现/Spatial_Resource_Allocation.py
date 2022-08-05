"""
论文复现：Spatial Resource Allocation for Emerging Epidemics: A Comparison of Greedy , Myopic, and Dynamic Policies

开始日期：2022-4-14 / 完成日期：2022-5-1

author：@陈典
"""
from openpyxl import load_workbook
from SEAQURD_NEW import SEAQURD
from Epidemic_Model_Final import Epidemic_Model
from LP import LP
import math
import numpy as np


class Spatial_Resource_Allocation:
    def __init__(self,file_name):
        self.file_name = file_name
        self.book = load_workbook(file_name)  # 加载文件
        self.region_num = 3
        self.sheet = []  # 加载每个地区（按顺序：0，1，2......）
        for i in range(self.region_num):
            self.sheet.append(self.book[f'{i}'])  # 将每一个sheet存入列表

        # 总人数（不变）
        self.total_population = []
        for i in range(self.region_num):
            self.total_population.append(self.sheet[i].cell(1, 2).value)

        self.dist = self.dist_type_1()  # 距离倒数矩阵

        self.start = 0  # 开始时间点
        self.end = 27  # 结束时间点（20.4.1-21.1.1，此处为275）
        self.t_num = self.end - self.start + 1  # 时间长度

        self.predict_num = 28  # 预测未来天数

        self.ksi = 0.5  # 论文给定的柯西值
        self.a = np.zeros((self.predict_num, self.region_num))  # 决策病床数（Myopic_LP）
        self.a_2 = np.zeros((self.predict_num, self.region_num))  # 决策病床数（Benchmark）
        self.b = [200 * (i + 1) for i in range(20)]  # 每周增加的病床数（多种情况，用于对比）
        self.r = 0.1  # 最大部署率

        self.simulate_para_list = [[] for i in range(self.predict_num)]  # 存放预测仿真期每一天的参数值
        self.simulate_result = np.zeros((self.region_num, self.predict_num))  # 存放自然状态仿真结果（对照组）
        self.simulate_a_result = np.zeros((self.region_num, self.predict_num))  # 存放Myopic_LP的仿真结果
        self.simulate_b_result = np.zeros((self.region_num, self.predict_num))  # 存放Benchmark的仿真结果

        self.pure_addition = 0  # 存放不安排病房的未来四周新增（纯新增潜伏者）
        self.a_addition = {}  # 存放Myopic_LP算法下，安排病房的未来四周新增（对应b的不同情况）
        for b in self.b:
            self.a_addition[f'b={b}'] = 0
        self.b_addition = {}  # 存放Benchmark算法下，安排病房的未来四周新增（对应b的不同情况）
        for b in self.b:
            self.b_addition[f'b={b}'] = 0

        self.algorithm()

    def algorithm(self):
        # （1）用原始传染病模型确定各仿真节点对应参数 + 仿真结果
        simulation = Epidemic_Model(self.file_name)
        self.simulate_para_list = simulation.simulate_para_list

        # 记录拟合，用于取出最后一个点当做起点
        S_fit, E_fit, A_fit, Q_fit = simulation.S, simulation.E, simulation.A, simulation.Q
        U_fit, R_fit, D_fit = simulation.U, simulation.R, simulation.D

        # 无病床安排（自然状态-对照组）：
        print('开始执行自然状态（对照组）')
        S_0_pre = [0 for i in range(self.region_num)]  # 滚动预测起点（为拟合的最后一天）
        E_0_pre = [0 for i in range(self.region_num)]
        A_0_pre = [0 for i in range(self.region_num)]
        Q_0_pre = [0 for i in range(self.region_num)]
        U_0_pre = [0 for i in range(self.region_num)]
        R_0_pre = [0 for i in range(self.region_num)]
        D_0_pre = [0 for i in range(self.region_num)]
        for i in range(self.region_num):
            S_0_pre[i] = S_fit[i][-1]
            E_0_pre[i] = E_fit[i][-1]
            A_0_pre[i] = A_fit[i][-1]
            Q_0_pre[i] = Q_fit[i][-1]
            U_0_pre[i] = U_fit[i][-1]
            R_0_pre[i] = R_fit[i][-1]
            D_0_pre[i] = D_fit[i][-1]

        start = self.end

        a_0 = [0 for i in range(self.region_num)]

        for t in range(self.predict_num):

            simulation_nature = SEAQURD(self.region_num, self.file_name, start, start + 1, self.total_population,
                                        S_0_pre, E_0_pre, A_0_pre, Q_0_pre, U_0_pre, R_0_pre, D_0_pre,
                                        self.simulate_para_list[t], self.ksi, a_0)
            for i in range(self.region_num):
                self.simulate_result[i][t] = simulation_nature.E[i][-1]

            # 更新起点（即向前推一天）
            start += 1
            for i in range(self.region_num):
                S_0_pre[i] = simulation_nature.S[i][-1]
                E_0_pre[i] = simulation_nature.E[i][-1]
                A_0_pre[i] = simulation_nature.A[i][-1]
                Q_0_pre[i] = simulation_nature.Q[i][-1]
                U_0_pre[i] = simulation_nature.U[i][-1]
                R_0_pre[i] = simulation_nature.R[i][-1]
                D_0_pre[i] = simulation_nature.D[i][-1]

        # Myopic_LP算法策略：
        print('开始执行Myopic_LP算法策略')
        for b in self.b:
            print(f'b（每周新增病床数）：{b}')
            S_0_pre = [0 for i in range(self.region_num)]  # 滚动预测起点（为拟合的最后一天）
            E_0_pre = [0 for i in range(self.region_num)]
            A_0_pre = [0 for i in range(self.region_num)]
            Q_0_pre = [0 for i in range(self.region_num)]
            U_0_pre = [0 for i in range(self.region_num)]
            R_0_pre = [0 for i in range(self.region_num)]
            D_0_pre = [0 for i in range(self.region_num)]
            for i in range(self.region_num):
                S_0_pre[i] = S_fit[i][-1]
                E_0_pre[i] = E_fit[i][-1]
                A_0_pre[i] = A_fit[i][-1]
                Q_0_pre[i] = Q_fit[i][-1]
                U_0_pre[i] = U_fit[i][-1]
                R_0_pre[i] = R_fit[i][-1]
                D_0_pre[i] = D_fit[i][-1]

            start = self.end

            B_before = 0
            a_before = [0 for i in range(self.region_num)]

            for t in range(self.predict_num):  # （2）对各仿真节点进行迭代规划
                if t % 7 == 0:
                    b_t = b
                else:
                    b_t = 0
                print(f'迭代仿真第 {t + 1} / {self.predict_num} 天')

                # （2-1）混合整数规划模型，确定最佳的a值
                model = LP(self.region_num, self.total_population, B_before, a_before, b_t, b * self.r, self.ksi,
                           self.simulate_para_list[t], self.dist, S_0_pre, A_0_pre, U_0_pre)

                self.a[t] = model.a  # 写入a值

                # （2-2）代入a值，利用new传染病模型，仿真预测出下一期
                simulation_a = SEAQURD(self.region_num, self.file_name, start, start + 1, self.total_population,
                                       S_0_pre, E_0_pre, A_0_pre, Q_0_pre, U_0_pre, R_0_pre, D_0_pre,
                                       self.simulate_para_list[t], self.ksi, self.a[t])
                for i in range(self.region_num):  # 存入结果
                    # self.simulate_a_result[i][t] = simulation_a.I[i][-1] - simulation_a.Q[i][-1]  # Q是已经入院的
                    self.simulate_a_result[i][t] = simulation_a.E[i][-1]  # Q是已经入院的

                # 更新起点（即向前推进一天）
                start += 1
                for i in range(self.region_num):
                    S_0_pre[i] = simulation_a.S[i][-1]
                    E_0_pre[i] = simulation_a.E[i][-1]
                    A_0_pre[i] = simulation_a.A[i][-1]
                    Q_0_pre[i] = simulation_a.Q[i][-1]
                    U_0_pre[i] = simulation_a.U[i][-1]
                    R_0_pre[i] = simulation_a.R[i][-1]
                    D_0_pre[i] = simulation_a.D[i][-1]

                # 更新B_before和a_before
                B_before = model.B
                a_before = model.a

            for t in range(1, self.predict_num):
                for i in range(self.region_num):  # 对净增长人数进行逆运算
                    self.a_addition[f'b={b}'] += self.simulate_a_result[i][t] - self.simulate_a_result[i][t - 1] \
                                                 + self.simulate_para_list[t][3][i] * self.simulate_a_result[i][t - 1]

        # Benchmark算法策略：
        print('开始执行Benchmark算法策略')
        for b in self.b:
            print(f'b（每周新增病床数）：{b}')
            S_0_pre = [0 for i in range(self.region_num)]  # 滚动预测起点（为拟合的最后一天）
            E_0_pre = [0 for i in range(self.region_num)]
            A_0_pre = [0 for i in range(self.region_num)]
            Q_0_pre = [0 for i in range(self.region_num)]
            U_0_pre = [0 for i in range(self.region_num)]
            R_0_pre = [0 for i in range(self.region_num)]
            D_0_pre = [0 for i in range(self.region_num)]
            for i in range(self.region_num):
                S_0_pre[i] = S_fit[i][-1]
                E_0_pre[i] = E_fit[i][-1]
                A_0_pre[i] = A_fit[i][-1]
                Q_0_pre[i] = Q_fit[i][-1]
                U_0_pre[i] = U_fit[i][-1]
                R_0_pre[i] = R_fit[i][-1]
                D_0_pre[i] = D_fit[i][-1]

            start = self.end

            for t in range(self.predict_num):  # （2）对各仿真节点进行迭代规划
                print(f'迭代仿真第 {t + 1} / {self.predict_num} 天')

                if t % 7 == 0:
                    a_2 = [0 for i in range(self.region_num)]  # 对本周新增资源进行比例分配
                    for i in range(self.region_num):
                        total_E = sum(E_0_pre)
                        a_2[i] = round(E_0_pre[i] / total_E * b)  # 四舍五入取整，按比例分配
                    # 每次都从头加即可（因为是累加的嘛），这样做的话，在非分配时间点就可以只计算人数即可
                    for tt in range(t, self.predict_num):
                        self.a_2[tt] = a_2  # （好像是+=是对的）

                simulation_b = SEAQURD(self.region_num, self.file_name, start, start + 1, self.total_population,
                                       S_0_pre, E_0_pre, A_0_pre, Q_0_pre, U_0_pre, R_0_pre, D_0_pre,
                                       self.simulate_para_list[t], self.ksi, self.a_2[t])
                for i in range(self.region_num):  # 存入结果
                    # self.simulate_a_result[i][t] = simulation_a.I[i][-1] - simulation_a.Q[i][-1]  # Q是已经入院的
                    self.simulate_b_result[i][t] = simulation_b.E[i][-1]

                # 更新起点（即向前推进一天）
                start += 1
                for i in range(self.region_num):
                    S_0_pre[i] = simulation_b.S[i][-1]
                    E_0_pre[i] = simulation_b.E[i][-1]
                    A_0_pre[i] = simulation_b.A[i][-1]
                    Q_0_pre[i] = simulation_b.Q[i][-1]
                    U_0_pre[i] = simulation_b.U[i][-1]
                    R_0_pre[i] = simulation_b.R[i][-1]
                    D_0_pre[i] = simulation_b.D[i][-1]

            for t in range(1, self.predict_num):
                for i in range(self.region_num):  # 对净增长人数进行逆运算
                    self.b_addition[f'b={b}'] += self.simulate_b_result[i][t] - self.simulate_b_result[i][t - 1] \
                                                 + self.simulate_para_list[t][3][i] * self.simulate_b_result[i][t - 1]

        # （3）最后，对比新增
        for t in range(1, self.predict_num):
            for i in range(self.region_num):  # 对净增长人数进行逆运算（自然状态）
                self.pure_addition += self.simulate_result[i][t] - self.simulate_result[i][t - 1] \
                                      + self.simulate_para_list[t][3][i] * self.simulate_result[i][t - 1]

        print(f'训练集：{(self.end - self.start + 1) / 7}周')
        print(f'自然状态下，未来四周潜伏者净增长数量：')
        print(self.pure_addition)
        print('---------------------------------------')

        print(f'Myopic_LP策略下，未来四周潜伏者净增长数量：')
        print(self.a_addition)
        print(f'病床分配结果：')
        print(self.a)
        print('---------------------------------------')
        print(f'Benchmark策略下，未来四周潜伏者净增长数量：')
        print(self.b_addition)
        print(f'病床分配结果：')
        print(self.a_2)
        print('---------------------------------------')

    def dist_type_1(self):
        dist = np.zeros((self.region_num, self.region_num))

        for i in range(self.region_num):
            for j in range(self.region_num):
                if i != j:
                    dist[i][j] = 1 / math.sqrt((self.sheet[i].cell(1,4).value - self.sheet[j].cell(1,4).value) ** 2 +
                                               (self.sheet[i].cell(1,5).value - self.sheet[j].cell(1,5).value) ** 2)

        return dist


if __name__ == '__main__':
    Spatial_Resource_Allocation('American_data.xlsx')

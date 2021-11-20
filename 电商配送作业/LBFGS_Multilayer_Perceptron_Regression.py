import numpy as np
import matplotlib.pyplot as plt
from sklearn.neural_network import MLPRegressor
from sklearn.metrics import mean_squared_error
from openpyxl import load_workbook
'''神经网络-多层感知机回归器：done'''


class LBFGS_MLP_Regression:  # 神经网络-多层感知机回归器
    def __init__(self,file_name):
        self.book = load_workbook(file_name)  # 读取数据
        self.sheet = self.book['inventory2']
        self.num = 12  # 历史数据数量
        self.num_future = 0  # 需要预测的未来天数

        self.x = []  # 初始化训练集x
        self.y = []  # 初始化训练集y，list形式即可
        # 同理：需要预测的测试集
        # self.x_test = [i for i in range(self.num+1,self.num+1+self.num_future)]
        self.x_test = []
        self.y_test = []

        for i in range(self.num):
            self.x.append([i+1,self.sheet.cell(6,2+i).value])

        for i in range(self.num):  # 录入表格数据
            self.y.append(self.sheet.cell(1,2+i).value)

        for i in range(self.num_future):
            self.y_test.append(self.sheet.cell(1,2+self.num+i).value)
            self.x_test.append([self.num+1+i,self.sheet.cell(6,2+self.num+i).value])

        self.regression()

    def regression(self):  # 回归算法主体

        x_train = np.array(self.x).reshape((self.num,2))
        y_train = np.array(self.y).ravel()

        # x_test = np.array(self.x_test).reshape((self.num_future,2))

        # 注意：此时 x_train、x_test 的格式变成了array，列为特征（即多项式各项），行为样本
        # y_train不用管，[y,y,y,y,y,y,y,y,y]这种形式就可以

        # solver：权重优化求解器，对于小型数据集，“ lbfgs”可以收敛得更快并且性能更好
        # hidden_layer_sizes：第i个元素代表第i个隐藏层中的神经元数量
        # 小数据集不要用默认的“adam”
        reg = MLPRegressor(hidden_layer_sizes=(5,5), solver='lbfgs', max_iter=10000)
        reg.fit(x_train, y_train)  # 用训练集训练

        y_predict_train = reg.predict(x_train)  # 对训练集拟合一下，导出y值
        mse_train = mean_squared_error(y_train, y_predict_train)  # 11211565

        # y_predict_future = reg.predict(x_test)  # 对测试集拟合一下，导出y值

        # 画图
        tem_train = [i + 1 for i in range(self.num)]
        # tem_future = [i for i in range(self.num+1,self.num+self.num_future+1)]
        tem_all = [i + 1 for i in range(self.num + self.num_future)]

        plt.plot(tem_all, self.y + self.y_test, color='lightskyblue', label='Actual_curve', marker='.')
        plt.plot(tem_train, y_predict_train, color='red', label='Fitting_curve', marker='.')
        # plt.plot(tem_future, y_predict_future, color='palegreen', label='predict curve', marker='.')

        plt.text(0.6, 18000, f'MSE of Multilayer_Perceptron_Regression:\n\n{mse_train}', fontsize=15,
                 bbox={'boxstyle': 'round', 'facecolor': 'lightskyblue', 'alpha': 0.5})

        for i in range(len(tem_all)):
            if y_train[i] >= y_predict_train[i]:
                plt.plot([i + 1, i + 1], [0, y_train[i]], linestyle='--', color='lightgrey')
            else:
                plt.plot([i + 1, i + 1], [0, y_predict_train[i]], linestyle='--', color='lightgrey')

        plt.xlabel('Months (1~12)', fontsize=15)
        plt.ylabel('Inventory number', fontsize=15)
        plt.ylim([-1000, 24000])
        plt.legend(loc='upper left', fontsize=15, facecolor='lightyellow')
        plt.show()


if __name__ == '__main__':
    LBFGS_MLP_Regression('Shipment_test.xlsx')

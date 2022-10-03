"""
用于统计每一层BOM的：采购、外协、厂内协同的数量（共六层）
"""
from openpyxl import load_workbook


class BOM:
    def __init__(self,file_name):
        self.file_name = file_name
        self.book = load_workbook(self.file_name)
        self.sheet = self.book['基础']

        self.levels = []

        for i in range(7):
            self.levels.append({'BOM层级':i, '采购数量':0, '外协':0, '重装厂自制':0, '厂内协同':0
                                , '串行式协同':0, '穿插式协同':0})

        self.count()
        self.display()

    def count(self):

        for raw in range(2,1176):
            # 采购统计
            if str(self.sheet.cell(raw, 14).value) == 'BUY':

                levels = str(self.sheet.cell(raw, 2).value).count('-') + 1
                self.levels[levels]['采购数量'] += int(float(self.sheet.cell(raw, 8).value))

            # 外协、重装厂自制、厂内协同统计
            if str(self.sheet.cell(raw, 14).value) == 'MAK':

                if str(self.sheet.cell(raw, 25).value) == '整体外协':

                    levels = str(self.sheet.cell(raw, 2).value).count('-') + 1
                    self.levels[levels]['外协'] += int(float(self.sheet.cell(raw, 8).value))

                elif str(self.sheet.cell(raw, 24).value) == '33 重型装备厂':

                    levels = str(self.sheet.cell(raw, 2).value).count('-') + 1
                    self.levels[levels]['重装厂自制'] += int(float(self.sheet.cell(raw, 8).value))

                    tem = ''.join(i for i in str(self.sheet.cell(raw, 17).value) if not i.isdigit())  # 去掉数字
                    tem = tem.split('-')  # 分割步骤
                    check = []
                    for string in tem:
                        # print(tem, string, check)
                        if string in check:
                            self.levels[levels]['穿插式协同'] += int(float(self.sheet.cell(raw, 8).value))
                            break
                        else:
                            check.append(string)
                    if len(tem) == len(check):
                        self.levels[levels]['串行式协同'] += int(float(self.sheet.cell(raw, 8).value))

                    if '外协' in str(self.sheet.cell(raw, 17).value):
                        self.levels[levels]['外协'] += int(float(self.sheet.cell(raw, 8).value))
                    elif ('大热' in str(self.sheet.cell(raw, 17).value)) or ('大热' in str(self.sheet.cell(raw, 17).value)):
                        self.levels[levels]['厂内协同'] += int(float(self.sheet.cell(raw, 8).value))

                else:

                    levels = str(self.sheet.cell(raw, 2).value).count('-') + 1
                    self.levels[levels]['厂内协同'] += int(float(self.sheet.cell(raw, 8).value))

    def display(self):
        for i in range(len(self.levels)):
            print(self.levels[i])


if __name__ == '__main__':
    BOM('生产BOM.xlsx')

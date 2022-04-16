import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import random
import math
import time
from Region import Region
'''S：易感者、E：潜伏者、I：感染者、R：治愈者'''


class Metropolis_Hastings:  # Metropolis_Hastings采样算法
    # 输入：文件名，起点，终点，总人数，各群体人数初值
    def __init__(self,region_num,file_name,start,end,total_population,S,E,A,Q,U,R,D):
        self.book = load_workbook(file_name)
        self.sheet = []
        self.region_num = region_num
        for i in range(self.region_num):
            self.sheet.append(self.book[f'{i}'])  # 将每一个sheet存入列表

        self.start = start  # 开始时间点
        self.end = end  # 结束时间点
        self.T = self.end - self.start + 1  # 时间长度

        self.total_population = total_population  # 各州总人口数

        self.region = []  # 用于存放地区

        self.actual_ratio_list = np.zeros((self.region_num, self.T))  # 感染者比例系数（用于平衡人口对各区域拟合误差的影响）

        self.actual = [[] for i in range(self.region_num)]
        for i in range(self.region_num):
            for j in range(self.T):
                self.actual[i].append(self.sheet[i].cell(1, self.start+j+89).value)
        # print(self.actual)

        for i in range(self.region_num):
            self.region.append(Region(f'{self.sheet[i].cell(1, 1).value}',
                                      [self.sheet[i].cell(1, 4).value, self.sheet[i].cell(1, 5).value],
                                      self.total_population[i],
                                      S[i], E[i], A[i], Q[i], U[i], R[i], D[i]))

        self.m = np.zeros((len(self.region), len(self.region)))  # 初始化距离矩阵

        self.I = np.zeros((len(self.region), self.T))  # 即：所有感染者数量 = A + Q + U

        # 两种生成距离倒数矩阵的方式：1：按照坐标计算，2：直接指定
        self.dist = self.dist_type_1()
        # self.dist = self.dist_type_2()

        self.iter = 50000  # 采样算法迭代次数（次数——收敛？？？？）

        self.para = []  # 用于存放最终拟合好的参数集合

        print('采样算法：读取数据完毕')

        self.observe_para = [[[]for p in range(10)] for i in range(self.region_num)]  # 查看参数的收敛情况

        self.sampling()

    def sampling(self):
        # 计算各区域感染者比例系数
        self.actual_ratio()

        # 参数赋初值
        beta_e = [0.04 for i in range(self.region_num)]
        beta_a = [0.04 for i in range(self.region_num)]
        beta_u = [0.04 for i in range(self.region_num)]
        alpha = [0.2 for i in range(self.region_num)]
        delta_a = [0.00015 for i in range(self.region_num)]
        delta_q = [0.00015 for i in range(self.region_num)]
        delta_u = [0.00015 for i in range(self.region_num)]
        gamma_a = [0.013 for i in range(self.region_num)]
        gamma_q = [0.013 for i in range(self.region_num)]
        gamma_u = [0.013 for i in range(self.region_num)]

        p = 0.3
        q = 0.6
        c_0 = 0.99

        # 初始化个群体list
        # 每行对应一个地区，列为时间间隔
        S = np.zeros((len(self.region), self.T))
        E = np.zeros((len(self.region), self.T))
        A = np.zeros((len(self.region), self.T))
        Q = np.zeros((len(self.region), self.T))
        U = np.zeros((len(self.region), self.T))
        R = np.zeros((len(self.region), self.T))
        D = np.zeros((len(self.region), self.T))
        for i in range(len(self.region)):
            S[i][0] = self.region[i].S
            E[i][0] = self.region[i].E
            A[i][0] = self.region[i].A
            Q[i][0] = self.region[i].Q
            U[i][0] = self.region[i].U
            R[i][0] = self.region[i].R
            D[i][0] = self.region[i].D

        # （1）计算初始的SSE值
        # self.m_calculate(c_0)  # 放在这里，代表c_0固定不变，放在algorithm函数开头，代表也需要拟合
        SSE = self.algorithm(S,E,A,Q,U,R,D,beta_e,beta_a,beta_u,alpha,delta_a,delta_q,delta_u,
                             gamma_a,gamma_q,gamma_u,p,q,c_0)

        # （2）迭代采样
        sample_start_t = time.time()
        for i in range(self.iter):
            if i % 5000 == 0:
                time_tem = time.time()
                print(f'完成迭代{i}次，当前用时{time_tem - sample_start_t}s')

            # 查看参数的收敛情况（每迭代50次记录一次）
            for r in range(self.region_num):
                self.observe_para[r][0].append(beta_e[r])
                self.observe_para[r][1].append(beta_a[r])
                self.observe_para[r][2].append(beta_u[r])
                self.observe_para[r][3].append(alpha[r])
                self.observe_para[r][4].append(delta_a[r])
                self.observe_para[r][5].append(delta_q[r])
                self.observe_para[r][6].append(delta_u[r])
                self.observe_para[r][7].append(gamma_a[r])
                self.observe_para[r][8].append(gamma_q[r])
                self.observe_para[r][9].append(gamma_u[r])

            # （2-1）记录上一次状态
            SSE_bef = SSE  # 记录上一次SSE结果

            beta_e_bef = beta_e  # 记录上一次参数结果
            beta_a_bef = beta_a
            beta_u_bef = beta_u
            alpha_bef = alpha
            delta_a_bef = delta_a
            delta_q_bef = delta_q
            delta_u_bef = delta_u
            gamma_a_bef = gamma_a
            gamma_q_bef = gamma_q
            gamma_u_bef = gamma_u
            p_bef = p
            q_bef = q
            c_0_bef = c_0

            # （2-2）产生新解，即：从均匀分布中随机抽取，并计算新的SSE
            beta_e = [random.uniform(0.02, 1) for i in range(self.region_num)]
            beta_a = [random.uniform(0.02, 1) for i in range(self.region_num)]
            beta_u = [random.uniform(0.02, 1) for i in range(self.region_num)]
            alpha = [random.uniform(0.15, 0.25) for i in range(self.region_num)]
            delta_a = [random.uniform(0, 0.005) for i in range(self.region_num)]
            delta_q = [random.uniform(0, 0.005) for i in range(self.region_num)]
            delta_u = [random.uniform(0, 0.005) for i in range(self.region_num)]
            gamma_a = [random.uniform(0, 0.3) for i in range(self.region_num)]
            gamma_q = [random.uniform(0, 0.3) for i in range(self.region_num)]
            gamma_u = [random.uniform(0, 0.3) for i in range(self.region_num)]
            p = random.uniform(0.1,0.5)
            q = random.uniform(0.5,0.8)
            c_0 = random.uniform(0.96,1)

            SSE = self.algorithm(S, E, A, Q, U, R, D, beta_e, beta_a, beta_u, alpha, delta_a, delta_q, delta_u,
                                 gamma_a, gamma_q, gamma_u, p, q, c_0)

            # （2-3）计算LR值，以一定概率接受SSE更小的结果
            LR = math.exp(min(700, (-SSE + SSE_bef)))  # 700是因为限制一下math.exp上限，防止报错
            r = random.uniform(0, 1)
            if LR >= r:
                continue
            else:
                beta_e = beta_e_bef  # 若接受，更新参数结果，更新SSE结果
                beta_a = beta_a_bef
                beta_u = beta_u_bef
                alpha = alpha_bef
                delta_a = delta_a_bef
                delta_q = delta_q_bef
                delta_u = delta_u_bef
                gamma_a = gamma_a_bef
                gamma_q = gamma_q_bef
                gamma_u = gamma_u_bef
                p = p_bef
                q = q_bef
                c_0 = c_0_bef

                SSE = SSE_bef
        sample_end_t = time.time()
        print(f'采样迭代总时间：{sample_end_t-sample_start_t}s')
        # （3）存入最终拟合好的参数
        self.para = [beta_e,beta_a,beta_u,alpha,delta_a,delta_q,delta_u,gamma_a,gamma_q,gamma_u,p,q,c_0]
        print('SSE:',SSE)

    def algorithm(self,S,E,A,Q,U,R,D,beta_e,beta_a,beta_u,alpha,delta_a,delta_q,delta_u,gamma_a,gamma_q,gamma_u,p,q,c_0):
        self.m_calculate(c_0)  # 放在algorithm函数开头，代表也需要拟合
        for i in range(len(self.region)):

            for t in range(1, self.T):

                e_rate = 0
                a_rate = 0
                u_rate = 0
                for j in range(len(self.region)):
                    e_rate += E[j][t - 1] / self.region[j].N * beta_e[j] * self.m[i][j]
                    a_rate += A[j][t - 1] / self.region[j].N * beta_a[j] * self.m[i][j]
                    u_rate += U[j][t - 1] / self.region[j].N * beta_u[j] * self.m[i][j]

                S[i][t] = S[i][t - 1] - S[i][t - 1] * (e_rate + a_rate + u_rate)
                E[i][t] = E[i][t - 1] + S[i][t - 1] * (e_rate + a_rate + u_rate) - alpha[i] * E[i][t - 1]
                A[i][t] = A[i][t - 1] + p * alpha[i] * E[i][t - 1] - delta_a[i] * A[i][t - 1] - gamma_a[i] * A[i][t - 1]
                Q[i][t] = Q[i][t - 1] + (1 - p) * q * alpha[i] * E[i][t - 1] \
                          - delta_q[i] * Q[i][t - 1] - gamma_q[i] * Q[i][t - 1]
                U[i][t] = U[i][t - 1] + (1 - p) * (1 - q) * alpha[i] * E[i][t - 1] \
                          - delta_u[i] * U[i][t - 1] - gamma_u[i] * U[i][t - 1]
                R[i][t] = R[i][t - 1] + gamma_a[i] * A[i][t - 1] \
                          + gamma_q[i] * Q[i][t - 1] \
                          + gamma_u[i] * U[i][t - 1]
                D[i][t] = D[i][t - 1] + delta_a[i] * A[i][t - 1] \
                          + delta_q[i] * Q[i][t - 1] \
                          + delta_u[i] * U[i][t - 1]
        I = A + Q + U
        SSE = 0

        for i in range(self.T):  # SSE计算方法：对各区域的各感染人数最小二乘（且乘以感染者比例系数去除人口影响）
            for j in range(self.region_num):
                SSE += ((I[j][i] - self.actual[j][i]) * self.actual_ratio_list[j][i]) ** 2

        return SSE

    def m_calculate(self,c_0):
        for i in range(len(self.region)):
            for j in range(len(self.region)):
                if i == j:
                    self.m[i][j] = c_0
                else:
                    self.m[i][j] = (1 - c_0) * self.dist[i][j] / sum(self.dist[:, j])

    def actual_ratio(self):
        total_actual = []
        for i in range(self.T):
            act = 0
            for j in range(self.region_num):
                act += self.actual[j][i]
            total_actual.append(act)

        for i in range(self.T):
            for j in range(self.region_num):
                self.actual_ratio_list[j][i] = total_actual[i] / self.actual[j][i] / self.region_num

    def dist_type_1(self):
        dist = np.zeros((len(self.region), len(self.region)))

        for i in range(len(self.region)):
            for j in range(len(self.region)):
                if i != j:
                    dist[i][j] = 1 / math.sqrt((self.region[i].location[0] - self.region[j].location[0]) ** 2 +
                                               (self.region[i].location[1] - self.region[j].location[1]) ** 2)

        return dist

    def dist_type_2(self):
        dist = np.zeros((len(self.region), len(self.region)))

        dist[0][1] = dist[1][0] = 1 / 38
        dist[0][2] = dist[2][0] = 1 / 18
        dist[1][2] = dist[2][1] = 1 / 24

        return dist


if __name__ == '__main__':
    Metropolis_Hastings(6, 'American_data.xlsx', 0, 0 + 60 - 1,
                        [4908621, 734002, 7378494, 3038999, 39937489, 5845526],
                        [4908621 - 3870, 734002 - 281, 7378494 - 3705, 3038999 - 1374, 39937489 - 23403, 5845526 - 7551],
                        [0, 0, 0, 0, 0, 0],
                        [0, 0, 0, 0, 0, 0],
                        [0, 0, 0, 0, 0, 0],
                        [3870, 281, 3705, 1374, 23403, 7551],
                        [0, 0, 0, 0, 0, 0],
                        [0, 0, 0, 0, 0, 0])
